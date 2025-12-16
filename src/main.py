"""Основной модуль расчёта приростов по клиентам СПОД.

Вся бизнес-логика собрана в одном файле main.py согласно требованиям.
"""

from __future__ import annotations

import csv
import datetime as dt
import operator
import traceback
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Set, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


SettingsTree = Dict[str, Any]
SELECTED_MANAGER_ID_COL = "Таб. номер ВКО (выбранный)"
SELECTED_MANAGER_NAME_COL = "ВКО (выбранный)"
DIRECT_MANAGER_ID_COL = "Таб. номер ВКО (по файлу)"
DIRECT_MANAGER_NAME_COL = "ВКО (по файлу)"


def build_settings_tree() -> SettingsTree:
    """Возвращает вложенную структуру настроек проекта."""

    # Структура intentionally verbose:
    # - "files" описывает, какие книги и листы участвуют, и даёт словарь alias↔source.
    # - "filters" держит правила отбраковки и переиспользуется в drop_forbidden_rows.
    # - "defaults"/"identifiers" управляют «заглушками» и форматированием ID.
    # - "spod"/"contest" отвечают за выгрузку и коды турнира.
    # - "variants" задаёт листы Excel (кластеры клиентских ключей).

    return {
        "files": {
            # sheet — лист по умолчанию. Если у конкретного файла другой лист, задайте его в items.
            "sheet": "Sheet1",
            # items: для каждого XLSX указываем ключ (не менять без причины), ярлык периода и имя файла в каталоге IN.
            #  - Хотите обрабатывать другие файлы? Меняйте file_name или добавляйте новые записи по аналогии.
            #  - Нужно читать несколько листов? Задайте "sheet": "Имя листа" у соответствующего item.
            #  - Если file_name для T-2 пустое, используется логика с 2 файлами, иначе с 3 файлами.
            #  - columns и filters.drop_rules можно задать для каждого файла отдельно.
            #    Если columns или filters.drop_rules - пустые массивы [], используются значения из defaults.
            "items": [
                {
                    "key": "current",          # фиксированный ключ T-0; лучше не переименовывать.
                    "label": "T-0",            # подпись, которая пойдёт в логи.
                    "file_name": "2025_M-10.xlsx",
                    "sheet": "Sheet1",
                    # Колонки для этого файла (если пустой массив [], используются из defaults.columns)
                    "columns": [
                        # {"alias": "tb", "source": "Короткое ТБ"},
                        # {"alias": "gosb", "source": "Полное ГОСБ"},
                        # {"alias": "manager_name", "source": "ФИО"},
                        # {"alias": "manager_id", "source": "Табельный номер"},
                        # {"alias": "client_id", "source": "ИНН"},
                        # {"alias": "fact_value", "source": "Факт"},
                    ],
                    # Фильтры для этого файла (если drop_rules пустой массив [], используются из defaults.drop_rules)
                    "filters": {
                        "drop_rules": [
                            # {"alias": "manager_name", "values": ["-", "Серая зона"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "manager_id", "values": ["-", "Green_Zone", "Tech_Sib"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "client_id", "values": ["Report_id не определен", "0", "000000000000"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "gosb", "values": ["9999"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                        ],
                    },
                },
                {
                    "key": "previous",         # фиксированный ключ T-1.
                    "label": "T-1",
                    "file_name": "2025_M-9.xlsx",
                    "sheet": "Sheet1",
                    # Колонки для этого файла (если пустой массив [], используются из defaults.columns)
                    "columns": [
                        # {"alias": "tb", "source": "Короткое ТБ"},
                        # {"alias": "gosb", "source": "Полное ГОСБ"},
                        # {"alias": "manager_name", "source": "ФИО"},
                        # {"alias": "manager_id", "source": "Табельный номер"},
                        # {"alias": "client_id", "source": "ИНН"},
                        # {"alias": "fact_value", "source": "Факт"},
                    ],
                    # Фильтры для этого файла (если drop_rules пустой массив [], используются из defaults.drop_rules)
                    "filters": {
                        "drop_rules": [
                            # {"alias": "manager_name", "values": ["-", "Серая зона"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "manager_id", "values": ["-", "Green_Zone", "Tech_Sib"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "client_id", "values": ["Report_id не определен", "0", "000000000000"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "gosb", "values": ["9999"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                        ],
                    },
                },
                {
                    "key": "previous2",        # фиксированный ключ T-2.
                    "label": "T-2",
                    "file_name": "2025_M-8.xlsx",  # Если пустое "", используется логика с 2 файлами
                    "sheet": "Sheet1",
                    # Колонки для этого файла (если пустой массив [], используются из defaults.columns)
                    "columns": [
                        # {"alias": "tb", "source": "Короткое ТБ"},
                        # {"alias": "gosb", "source": "Полное ГОСБ"},
                        # {"alias": "manager_name", "source": "ФИО"},
                        # {"alias": "manager_id", "source": "Табельный номер"},
                        # {"alias": "client_id", "source": "ИНН"},
                        # {"alias": "fact_value", "source": "Факт"},
                    ],
                    # Фильтры для этого файла (если drop_rules пустой массив [], используются из defaults.drop_rules)
                    "filters": {
                        "drop_rules": [
                            # {"alias": "manager_name", "values": ["-", "Серая зона"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "manager_id", "values": ["-", "Green_Zone", "Tech_Sib"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "client_id", "values": ["Report_id не определен", "0", "000000000000"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                            # {"alias": "gosb", "values": ["9999"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                        ],
                    },
                },
                {
                    "key": "single",          # фиксированный ключ для одного файла (только для use_files_count="one")
                    "label": "Single",        # подпись, которая пойдёт в логи.
                    "file_name": "2025_M-10.xlsx",  # Имя файла в каталоге IN (например, "2025_M-10.xlsx")
                    "sheet": "Sheet1",
                    # Колонки для этого файла (НЕ используются значения из defaults.columns, только свои)
                    # Для варианта 1 файл будет свой набор колонок, не связанный с default
                    "columns": [
                        # {"alias": "tb", "source": "Короткое ТБ"},
                        # {"alias": "gosb", "source": "Полное ГОСБ"},
                        # {"alias": "manager_name", "source": "ФИО"},
                        # {"alias": "manager_id", "source": "Табельный номер"},
                        # {"alias": "client_id", "source": "ИНН"},
                        # {"alias": "fact_value", "source": "Факт"},
                    ],
                    # Фильтры для этого файла (НЕ используются значения из defaults.drop_rules, только свои)
                    # Для варианта 1 файл будет свой набор фильтров, не связанный с default
                    "filters": {
                        # drop_rules: правила удаления строк (только свои, не из defaults)
                        "drop_rules": [
                            # {"alias": "manager_name", "values": ["-", "Серая зона"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                        ],
                        # in_rules: правила включения строк (что ДОЛЖНО попасть в расчет)
                        #   - Строка попадает в расчет только если она проходит ВСЕ условия из in_rules (И)
                        #   - И при этом НЕ попадает под drop_rules (исключается из DROP)
                        #   - Формат: {"alias": "имя_колонки", "values": ["значение1", "значение2"], "condition": "in" или "not_in"}
                        #   - "in": значение должно быть в списке values
                        #   - "not_in": значение НЕ должно быть в списке values
                        "in_rules": [
                            # {"alias": "manager_id", "values": ["12345", "67890"], "condition": "in"},
                            # {"alias": "tb", "values": ["ЦА"], "condition": "not_in"},
                        ],
                    },
                },
                # Файлы для нового варианта расчета (use_files_count="new"): 24 файла (12 для 2025 и 12 для 2024)
                # Для каждого файла действуют параметры как для базового расчета (columns и drop_rules)
                # Если пусто - берем из defaults
                *[
                    {
                        "key": f"2025_M-{i+1:02d}",
                        "label": f"2025_M-{i+1:02d}",
                        "file_name": f"2025_M-{i+1:02d}.xlsx",  # Имя файла в каталоге IN
                        "sheet": "Sheet1",
                        "columns": [],  # Если пустой массив [], используются из defaults.columns
                        "filters": {
                            "drop_rules": [],  # Если пустой массив [], используются из defaults.drop_rules
                        },
                    }
                    for i in range(12)
                ],
                *[
                    {
                        "key": f"2024_M-{i+1:02d}",
                        "label": f"2024_M-{i+1:02d}",
                        "file_name": f"2024_M-{i+1:02d}.xlsx",  # Имя файла в каталоге IN
                        "sheet": "Sheet1",
                        "columns": [],  # Если пустой массив [], используются из defaults.columns
                        "filters": {
                            "drop_rules": [],  # Если пустой массив [], используются из defaults.drop_rules
                        },
                    }
                    for i in range(12)
                ],
            ],
        },
        "defaults": {
            # Заглушки менеджера, которые попадут в итог, если T-0/T-1 не дали значений.
            #  - Можно указывать реальные ФИО/табельные номера из справочника (строки).
            "manager_name": "Не найден КМ",
            "manager_id": "90000009",
            # columns: общие колонки по умолчанию (используются, если в items для файла columns пустой массив)
            #  - Чтобы подставить другое поле из Excel, достаточно изменить "source".
            #  - Чтобы добавить ещё колонку, расширьте список и пропишите alias (английское имя) + source (русский заголовок).
            "columns": [
                {"alias": "tb", "source": "Короткое ТБ"},
                {"alias": "gosb", "source": "Полное ГОСБ"},
                {"alias": "manager_name", "source": "ФИО"},
                {"alias": "manager_id", "source": "Табельный номер"},
                {"alias": "client_id", "source": "ИНН"},
                {"alias": "fact_value", "source": "Факт"},
            ],
            # drop_rules: общие правила фильтрации по умолчанию (используются, если в items для файла filters.drop_rules пустой массив)
            #  - Можно дополнять массив values новыми маркерами (например, ["-", "N/A", "Удалить"]).
            #  - alias должен совпадать с alias из блока columns.
            #
            # ПАРАМЕТРЫ УСЛОВНОГО УДАЛЕНИЯ:
            #
            # remove_unconditionally (bool, по умолчанию True):
            #   - True: удаляем строки с запрещенными значениями (с учетом условий, если они заданы)
            #   - False: НЕ удаляем строки вообще, правило игнорируется
            #
            # check_by_inn (bool, по умолчанию False):
            #   - True: перед удалением проверяем по ИНН (client_id)
            #     Если по этому ИНН есть другие строки с НЕзапрещенными значениями в этой колонке - строка НЕ удаляется
            #     Если по этому ИНН все строки имеют запрещенные значения - строка удаляется
            #   - False: проверка по ИНН не выполняется
            #
            # check_by_tn (bool, по умолчанию False):
            #   - True: перед удалением проверяем по ТН (manager_id)
            #     Если по этому ТН есть другие строки с НЕзапрещенными значениями в этой колонке - строка НЕ удаляется
            #     Если по этому ТН все строки имеют запрещенные значения - строка удаляется
            #   - False: проверка по ТН не выполняется
            #
            # КОМБИНАЦИИ ПАРАМЕТРОВ:
            #
            # 1. remove_unconditionally=True, check_by_inn=False, check_by_tn=False:
            #    → Удаляем ВСЕ строки с запрещенными значениями (старая логика, безусловное удаление)
            #
            # 2. remove_unconditionally=True, check_by_inn=True, check_by_tn=False:
            #    → Удаляем строки с запрещенными значениями, НО:
            #      - Если по ИНН есть строки с другими (незапрещенными) значениями → строка НЕ удаляется
            #      - Если по ИНН все строки имеют запрещенные значения → строка удаляется
            #
            # 3. remove_unconditionally=True, check_by_inn=False, check_by_tn=True:
            #    → Удаляем строки с запрещенными значениями, НО:
            #      - Если по ТН есть строки с другими (незапрещенными) значениями → строка НЕ удаляется
            #      - Если по ТН все строки имеют запрещенные значения → строка удаляется
            #
            # 4. remove_unconditionally=True, check_by_inn=True, check_by_tn=True:
            #    → Удаляем строки с запрещенными значениями, НО:
            #      - Если по ИНН ИЛИ по ТН есть строки с другими (незапрещенными) значениями → строка НЕ удаляется (логика ИЛИ)
            #      - Если и по ИНН, и по ТН все строки имеют запрещенные значения → строка удаляется
            #
            # 5. remove_unconditionally=False (любые значения check_by_inn и check_by_tn):
            #    → Строки НЕ удаляются, правило полностью игнорируется
            #
            # ПРИМЕРЫ:
            #
            # Пример 1: Удалить все строки с tb="ЦА" безусловно
            #   {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False}
            #
            # Пример 2: Удалить строки с tb="ЦА", но оставить если по ИНН есть другие ТБ
            #   {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": True, "check_by_tn": False}
            #   Если у клиента ИНН=123 есть строки: tb="ЦА" и tb="МСК" → строка с tb="ЦА" НЕ удаляется
            #   Если у клиента ИНН=456 есть только строки: tb="ЦА" → строка с tb="ЦА" удаляется
            #
            # Пример 3: Удалить строки с tb="ЦА", но оставить если по ТН есть другие ТБ
            #   {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": True}
            #   Если у менеджера ТН=001 есть строки: tb="ЦА" и tb="МСК" → строка с tb="ЦА" НЕ удаляется
            #   Если у менеджера ТН=002 есть только строки: tb="ЦА" → строка с tb="ЦА" удаляется
            #
            # Пример 4: Удалить строки с tb="ЦА", но оставить если по ИНН ИЛИ по ТН есть другие ТБ
            #   {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": True, "check_by_tn": True}
            #   Если по ИНН есть другие ТБ ИЛИ по ТН есть другие ТБ → строка НЕ удаляется
            #   Если и по ИНН, и по ТН все строки имеют tb="ЦА" → строка удаляется
            "drop_rules": [
                {"alias": "manager_name", "values": ["-", "Серая зона"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                {"alias": "manager_id", "values": ["-", "Green_Zone", "Tech_Sib"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                {"alias": "client_id", "values": ["Report_id не определен", "0", "000000000000"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                {"alias": "tb", "values": ["ЦА"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
                {"alias": "gosb", "values": ["9999"], "remove_unconditionally": True, "check_by_inn": False, "check_by_tn": False},
            ],
        },
        "identifiers": {
            # Форматирование ID: задаём символ заполнения (fill_char) и итоговую длину.
            #  - Например, замените на {"fill_char": " ", "total_length": 6}, если нужен пробел и длина 6.
            "manager_id": {"fill_char": "0", "total_length": 8},
            "client_id": {"fill_char": "0", "total_length": 12},
        },
        "spod": {
            # Глобальные параметры имени файлов/логов.
            "file_prefix": "YEAR_SPOD_Active_Rost_ost",
            "log_topic": "spod",
            # Варианты выгрузки СПОД
            "variants": [
                {
                    "name": "SPOD_SCENARIO",
                    "calc_sheet_name": "CALC_SCENARIO",
                    # source_type: определяет тип источника данных и влияет на формирование SPOD выгрузки
                    # Варианты:
                    #   "scenario_summary" - использует данные из SUMMARY_TN (основной расчет прироста)
                    #     - FACT_VALUE берется из value_column (обычно "Прирост")
                    #     - В Excel НЕ добавляются процентильные колонки (Обогнал_всего_кол, Обогнали_меня_всего_кол и т.д.)
                    #     - percentile_type из варианта процентиля НЕ используется
                    #     - Используется для выгрузки фактических значений прироста
                    #   "scenario_percentile" - использует данные из SUMMARY_TN с процентилями (PERCENTILE_TN)
                    #     - FACT_VALUE берется из percentile_type настроек процентиля:
                    #       * если percentile_type="above" → FACT_VALUE = Обогнал_всего_%
                    #       * если percentile_type="below" → FACT_VALUE = Обогнали_меня_всего_%
                    #     - В Excel добавляются процентильные колонки: Обогнал_всего_кол, Обогнали_меня_всего_кол, Равных_всего_кол, Всего_КМ_всего
                    #     - percentile_type берется из variants.active_percentile_variant
                    #     - Используется для выгрузки процентильных метрик
                    "source_type": "scenario_summary",
                    "value_column": "Прирост",
                    "fact_value_filter": "all",  # Фильтр для вывода в SPOD (только положительные приросты)
                    "plan_value": 0.0,
                    "priority": 1,
                    "contest_code": "YEAR_SPOD",
                    "tournament_code": "ACTIVE_ROST_OST",
                    "contest_date": "01/01/2025",
                    "include_in_csv": True,
                },
                {
                    "name": "SPOD_SCENARIO_PERCENTILE",
                    "calc_sheet_name": "CALC_SCENARIO_PERC",
                    # source_type: определяет тип источника данных и влияет на формирование SPOD выгрузки
                    # Варианты:
                    #   "scenario_summary" - использует данные из SUMMARY_TN (основной расчет прироста)
                    #     - FACT_VALUE берется из value_column (обычно "Прирост")
                    #     - В Excel НЕ добавляются процентильные колонки (Обогнал_всего_кол, Обогнали_меня_всего_кол и т.д.)
                    #     - percentile_type из варианта процентиля НЕ используется
                    #     - Используется для выгрузки фактических значений прироста
                    #   "scenario_percentile" - использует данные из SUMMARY_TN с процентилями (PERCENTILE_TN)
                    #     - FACT_VALUE берется из percentile_type настроек процентиля:
                    #       * если percentile_type="above" → FACT_VALUE = Обогнал_всего_%
                    #       * если percentile_type="below" → FACT_VALUE = Обогнали_меня_всего_%
                    #     - В Excel добавляются процентильные колонки: Обогнал_всего_кол, Обогнали_меня_всего_кол, Равных_всего_кол, Всего_КМ_всего
                    #     - percentile_type берется из variants.active_percentile_variant
                    #     - Используется для выгрузки процентильных метрик
                    "source_type": "scenario_percentile",
                    "value_column": "Обогнал_всего_%",  # Используется для сортировки и фильтрации (НЕ для FACT_VALUE)
                    # percentile_value_type для FACT_VALUE берется из percentile_type настроек процентиля (variants.percentile_calculation.percentile_type)
                    "fact_value_filter": "all",  # Фильтр для вывода в SPOD (все неотрицательные процентили)
                    "plan_value": 0.0,
                    "priority": 1,
                    "contest_code": "YEAR_SPOD_P",
                    "tournament_code": "ACTIVE_ROST_OST_P",
                    "contest_date": "01/01/2025",
                    "include_in_csv": True,
                },
            ],
        },
        "main_calculation": {
            # Параметры основного расчета прироста
            # use_files_count: количество файлов для расчета (текстовые значения)
            #   "one" - расчет по одному файлу (используются параметры из one_file)
            #   "two" - расчет по двум файлам (используются параметры из two_files)
            #   "three" - расчет по трем файлам (используются параметры из three_files)
            #   "new" - расчет по 24 файлам (12 для 2025 и 12 для 2024) - поиск новых клиентов
            "use_files_count": "new",  # "one", "two", "three" или "new"
            
            # one_file: параметры для расчета по одному файлу (только для use_files_count="one")
            # Примечание: file_name, sheet, columns и filters находятся в files.items с key="single"
            "one_file": {
                # calculation_type: тип расчета для одного файла
                #   "count" - количество строк (сделок) для каждого ТН
                #   "max" - максимальная сумма среди строк для каждого КМ
                "calculation_type": "count",  # "count" или "max"
            },
            
            # two_files: параметры для расчета по двум файлам (только для use_files_count="two")
            "two_files": {
                # key_mode: режим агрегации данных
                #   "manager" - агрегация по manager_id (табельному номеру), суммируем в каждом файле по КМ, затем разница
                #   "client" - агрегация по client_id (ИНН), КМ определяется на конец периода (T-0 → T-1)
                "key_mode": "client",  # "manager" или "client"
                # include_tb: учитывать ли ТБ при расчете (только для key_mode="client")
                #   True - расчет с учетом ТБ (клиент привязан к КМ в рамках ТБ)
                #   False - расчет без учета ТБ (клиент привязан к КМ глобально)
                "include_tb": False,  # True или False
            },
            
            # three_files: параметры для расчета по трем файлам (только для use_files_count="three")
            "three_files": {
                # key_mode: режим агрегации данных
                #   "manager" - агрегация по manager_id (табельному номеру), суммируем в каждом файле по КМ, затем разница
                #   "client" - агрегация по client_id (ИНН), КМ определяется на конец периода (T-0 → T-1 → T-2)
                "key_mode": "client",  # "manager" или "client"
                # include_tb: учитывать ли ТБ при расчете (только для key_mode="client")
                #   True - расчет с учетом ТБ (клиент привязан к КМ в рамках ТБ)
                #   False - расчет без учета ТБ (клиент привязан к КМ глобально)
                "include_tb": False,  # True или False
            },
            
            # new_files: параметры для расчета по 24 файлам (только для use_files_count="new")
            # Поиск новых клиентов: ИНН с суммой факта в 2025 > 0, но сумма факта в 2024 = 0 или его нет
            "new_files": {
                # key_mode: режим агрегации данных
                #   "manager" - агрегация по manager_id (табельному номеру)
                #   "client" - агрегация по client_id (ИНН), КМ определяется на конец периода
                "key_mode": "client",  # "manager" или "client"
                # include_tb: учитывать ли ТБ при расчете (только для key_mode="client")
                #   True - расчет с учетом ТБ (клиент привязан к КМ в рамках ТБ)
                #   False - расчет без учета ТБ (клиент привязан к КМ глобально)
                "include_tb": False,  # True или False
            },
        },
        "percentile_calculation": {
            # Параметры расчета процентиля (кто кого обогнал)
            # percentile_type: тип процентиля
            #   "above" - рассчитывается процент КМ с меньшим результатом (кого я обогнал, кто ниже меня)
            #   "below" - рассчитывается процент КМ с большим результатом (кто меня обогнал, кто выше меня)
            "percentile_type": "above",  # "above" или "below"
            # percentile_group_by: уровень группировки для расчета процентиля
            #   "all" - сравнение среди всех КМ
            #   "tb" - сравнение только среди КМ с тем же ТБ
            #   "gosb" - сравнение только среди КМ с тем же ГОСБ
            #   "tb_and_gosb" - сравнение только среди КМ с тем же ТБ и ГОСБ одновременно
            "percentile_group_by": "all",  # "all", "tb", "gosb" или "tb_and_gosb"
            # percentile_filter: фильтр для данных при расчете процентилей
            #   ">=0" - расчет только по неотрицательным значениям
            #   ">0" - расчет только по положительным значениям
            #   "all" - расчет по всем значениям
            "percentile_filter": ">0",  # ">=0", ">0", "all" и т.д.
        },
        "excel_formatting": {
            # Параметры форматирования Excel листов
            # column_width: настройки ширины колонок
            "column_width": {
                "min_width": 15,  # Минимальная ширина колонки в пунктах
                "max_width": 150,  # Максимальная ширина колонки в пунктах
                "auto_fit": True,  # Автоматическая подстройка ширины по содержимому
            },
            # wrap_text: включить перенос текста по строкам для всех ячеек
            "wrap_text": True,  # True - включить перенос текста, False - отключить
        },
        "report_layout": {
            # Управляет тем, какие листы попадают в основной Excel (пустой список = блок отключён).
            # По умолчанию записываются SUMMARY_TN (объединенный с процентилями) и SUMMARY_INN (для вариантов 2 и 3)
            "summary_sheets": ["SUMMARY_TN", "SUMMARY_INN"],
            "spod_variants": ["SPOD_SCENARIO", "SPOD_SCENARIO_PERCENTILE"],
            # raw_sheets — очищенные исходники T-0/T-1/T-2.
            "raw_sheets": ["RAW_T0", "RAW_T1", "RAW_T2"],
        },
    }


def build_column_profiles(columns: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    """Формирует маппинги alias↔source для переименования колонок."""

    # rename_map: перевод оригинальных колонок Excel в машинные имена;
    # alias_to_source: обратное отображение для вывода человекочитаемых заголовков.
    rename_map = {column["source"]: column["alias"] for column in columns}
    alias_to_source = {column["alias"]: column["source"] for column in columns}
    return {"rename_map": rename_map, "alias_to_source": alias_to_source}


def build_drop_rules(rule_items: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Возвращает словарь правил фильтрации по колонкам.
    
    Каждое правило содержит:
    - values: кортеж запрещенных значений
    - remove_unconditionally: убирать ли всегда (по умолчанию True)
    - check_by_inn: проверять ли по ИНН (по умолчанию False)
    - check_by_tn: проверять ли по ТН (по умолчанию False)
    
    Args:
        rule_items: Список правил из конфигурации
    
    Returns:
        Словарь {alias: {values: tuple, remove_unconditionally: bool, check_by_inn: bool, check_by_tn: bool}}
    """
    result = {}
    for rule in rule_items:
        alias = rule["alias"]
        result[alias] = {
            "values": tuple(rule["values"]),
            "remove_unconditionally": rule.get("remove_unconditionally", True),
            "check_by_inn": rule.get("check_by_inn", False),
            "check_by_tn": rule.get("check_by_tn", False),
        }
    return result


def get_file_meta(file_section: Dict[str, Any], key: str) -> Dict[str, Any]:
    """Ищет метаданные файла по ключу."""

    for item in file_section["items"]:
        if item["key"] == key:
            return item
    raise KeyError(f"Не найдена конфигурация файла '{key}'")


def resolve_sheet_name(file_section: Dict[str, Any], file_key: str) -> str:
    """Определяет имя листа для конкретного файла (или общее значение)."""

    meta = get_file_meta(file_section, file_key)
    return meta.get("sheet") or file_section.get("sheet", "Sheet1")


def get_file_columns(file_section: Dict[str, Any], file_key: str, defaults: Dict[str, Any], use_defaults: bool = True) -> List[Dict[str, str]]:
    """Возвращает колонки для конкретного файла.
    
    Логика:
    - Если в items для файла есть columns и он не пустой (не пустой массив), используется columns из items
    - Если в items для файла columns отсутствует или это пустой массив:
      - Если use_defaults=True, используется columns из defaults
      - Если use_defaults=False, возвращается пустой список
    
    Args:
        file_section: Секция files из настроек
        file_key: Ключ файла ("current", "previous", "previous2", "single")
        defaults: Секция defaults из настроек
        use_defaults: Использовать ли значения по умолчанию, если columns пустой (по умолчанию True)
    
    Returns:
        Список колонок для файла
    """
    meta = get_file_meta(file_section, file_key)
    if "columns" in meta and isinstance(meta["columns"], list) and len(meta["columns"]) > 0:
        return meta["columns"]
    if use_defaults:
        return defaults.get("columns", [])
    return []


def get_file_filters(file_section: Dict[str, Any], file_key: str, defaults: Dict[str, Any], use_defaults: bool = True) -> Dict[str, Any]:
    """Возвращает фильтры для конкретного файла.
    
    Логика:
    - Если в items для файла есть filters.drop_rules и он не пустой (не пустой массив), используется drop_rules из items
    - Если в items для файла filters отсутствует или filters.drop_rules пустой массив:
      - Если use_defaults=True, используется drop_rules из defaults
      - Если use_defaults=False, возвращается пустой словарь с пустыми drop_rules
    
    Args:
        file_section: Секция files из настроек
        file_key: Ключ файла ("current", "previous", "previous2", "single")
        defaults: Секция defaults из настроек
        use_defaults: Использовать ли значения по умолчанию, если drop_rules пустой (по умолчанию True)
    
    Returns:
        Словарь с фильтрами для файла (включая drop_rules и in_rules)
    """
    meta = get_file_meta(file_section, file_key)
    if "filters" in meta and isinstance(meta["filters"], dict):
        drop_rules = meta["filters"].get("drop_rules", [])
        if isinstance(drop_rules, list) and len(drop_rules) > 0:
            # Возвращаем фильтры из items (включая in_rules, если есть)
            result = {"drop_rules": drop_rules}
            if "in_rules" in meta["filters"]:
                result["in_rules"] = meta["filters"]["in_rules"]
            else:
                result["in_rules"] = []
            return result
    
    # Если drop_rules пустой или отсутствует
    if use_defaults:
        return {"drop_rules": defaults.get("drop_rules", []), "in_rules": []}
    else:
        return {"drop_rules": [], "in_rules": []}


def parse_contest_date(contest_date: str) -> str:
    """Возвращает дату турнира в формате ISO."""

    parsed = dt.datetime.strptime(contest_date, "%d/%m/%Y")
    return parsed.strftime("%Y-%m-%d")


def get_manager_columns(mode: str) -> Mapping[str, str]:
    """Возвращает имена колонок для выбранного режима назначения менеджера."""

    mapping = {
        "latest": {
            "id": "Таб. номер ВКО_Актуальный",
            "name": "ВКО_Актуальный",
        },
        "current_period": {
            "id": "Таб. номер ВКО_T0",
            "name": "ВКО_T0",
        },
        "previous_period": {
            "id": "Таб. номер ВКО_T1",
            "name": "ВКО_T1",
        },
    }
    if mode not in mapping:
        raise ValueError(
            "Недопустимое значение manager_mode. Используйте latest, current_period или previous_period."
        )
    return mapping[mode]


def ensure_directories(directories: Iterable[Path]) -> None:
    """Создаёт недостающие каталоги."""

    for directory in directories:
        directory.mkdir(parents=True, exist_ok=True)


def timestamp_suffix() -> str:
    """Формирует суффикс вида _YYYYMMDD_HH_MM."""

    return dt.datetime.now().strftime("_%Y%m%d_%H_%M")


def format_identifier(value: Any, total_length: int, fill_char: str) -> str:
    """Преобразует числовой идентификатор с лидирующими символами."""

    text = "" if value is None else str(value).strip()
    if not text:
        return text
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        return digits.rjust(total_length, fill_char)
    return text


def safe_to_float(value: Any) -> Optional[float]:
    """Безопасно приводит значение к float."""

    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def build_filter_mask(series: pd.Series, condition: str) -> pd.Series:
    """Возвращает булев маск для фильтрации значений по условию."""

    normalized = condition.strip().lower().replace(" ", "")
    if not normalized or normalized in ("all", "все"):
        return pd.Series(True, index=series.index)

    # Поддерживаются операторы сравнения и записи вида ">0", "<=1000", "==0", "!=5".
    # Часть после оператора парсится как число (точка или запятая).
    available_ops = {
        ">=": operator.ge,
        "<=": operator.le,
        "==": operator.eq,
        "!=": operator.ne,
        ">": operator.gt,
        "<": operator.lt,
        "=": operator.eq,
    }

    for token in ("<=", ">=", "==", "!=", ">", "<", "="):
        if normalized.startswith(token):
            threshold_text = normalized[len(token) :]
            try:
                threshold = float(threshold_text.replace(",", "."))
            except ValueError as error:
                raise ValueError(
                    f"Не удалось распознать значение фильтра '{condition}'."
                ) from error
            comparator = available_ops[token]
            return comparator(series, threshold)

    raise ValueError(
        "Фильтр FACT_VALUE должен начинаться с одного из операторов "
        "(>=, <=, >, <, ==, !=, = ) или быть 'all/все'."
    )


def _compute_percentile_pair(series: pd.Series) -> Tuple[pd.Series, pd.Series, pd.Series, pd.Series, pd.Series, pd.Series]:
    """Вспомогательная функция: возвращает (обогнал_%, обогнали_%, обогнал_кол, обогнали_кол, равных_кол, всего_кол) для серии."""

    if series.empty:
        empty = pd.Series(0.0, index=series.index)
        return empty, empty, empty, empty, empty, empty

    rank_min = series.rank(method="min", ascending=True)
    rank_max = series.rank(method="max", ascending=True)
    count_equal = rank_max - rank_min + 1
    count_less = rank_min - 1
    count_greater = len(series) - rank_max
    total_count = len(series)

    obognal = ((count_less + 0.5 * (count_equal - 1)) / total_count) * 100
    obognali = ((count_greater + 0.5 * (count_equal - 1)) / total_count) * 100

    return obognal, obognali, count_less, count_greater, count_equal - 1, pd.Series(total_count, index=series.index)


def append_percentile_columns(
    table: pd.DataFrame,
    *,
    value_column: str,
    tb_column: Optional[str] = None,
) -> pd.DataFrame:
    """Добавляет в таблицу колонки процентных рангов и абсолютных значений.
    
    Функция-обертка для PercentileCalculator.append_percentile_columns.
    Сохранена для обратной совместимости.
    
    Args:
        table: DataFrame с данными для расчета процентилей
        value_column: Имя колонки со значениями для расчета
        tb_column: Имя колонки с ТБ для группировки (может быть None)
    
    Returns:
        DataFrame с добавленными колонками процентилей
    """
    return PercentileCalculator.append_percentile_columns(table, value_column=value_column, tb_column=tb_column)


def build_scenario_keys(key_mode: str, include_tb: bool) -> List[str]:
    """Возвращает список колонок для ключа сценария."""

    mapping = {
        "client": ["client_id"],
        "manager": ["manager_id"],
    }
    if key_mode not in mapping:
        raise ValueError("key_mode должен быть client или manager")
    keys = list(mapping[key_mode])
    if include_tb:
        keys.append("tb")
    return keys


def _ensure_manager_identity(
    value: Any,
    *,
    default_value: str,
    identifiers: Mapping[str, Any],
) -> str:
    """Возвращает табельный номер с учётом обязательной длины."""

    if not value or str(value).strip() == "":
        return format_identifier(
            default_value,
            total_length=identifiers["manager_id"]["total_length"],
            fill_char=identifiers["manager_id"]["fill_char"],
        )
    return format_identifier(
        value,
        total_length=identifiers["manager_id"]["total_length"],
        fill_char=identifiers["manager_id"]["fill_char"],
    )


def build_assignment_table(
    variant_df: pd.DataFrame,
    *,
    key_columns: List[str],
    manager_assignment: str,
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
    scenario_name: str,
    manager_tb_mapping: Optional[pd.Series] = None,
) -> pd.DataFrame:
    """Возвращает таблицу назначений (ключ ↔ выбранный КМ). Всегда добавляет ТБ, определяя его по табельному номеру."""

    if manager_assignment not in {"latest", "per_file"}:
        raise ValueError("manager_assignment должен быть latest или per_file.")

    if variant_df.empty:
        columns = key_columns + [
            SELECTED_MANAGER_ID_COL,
            SELECTED_MANAGER_NAME_COL,
            "Источник",
            "Факт_T0",
            "Факт_T1",
            "Прирост",
        ]
        if "ТБ" not in columns and "tb" not in columns:
            columns.append("ТБ")
        return pd.DataFrame(columns=columns)

    if manager_assignment == "latest":
        assignments = variant_df[key_columns].copy()
        assignments[SELECTED_MANAGER_ID_COL] = variant_df["Таб. номер ВКО_Актуальный"]
        assignments[SELECTED_MANAGER_NAME_COL] = variant_df["ВКО_Актуальный"]
        assignments["Источник"] = "LATEST"
        assignments["Факт_T0"] = variant_df["Факт_T0"].fillna(0.0)
        assignments["Факт_T1"] = variant_df["Факт_T1"].fillna(0.0)
        assignments["Прирост"] = variant_df["Прирост"].fillna(0.0)
        
        # Добавляем ТБ, если его нет, определяя по табельному номеру
        if "ТБ" not in assignments.columns and "tb" not in assignments.columns:
            if manager_tb_mapping is not None:
                assignments["ТБ"] = assignments[SELECTED_MANAGER_ID_COL].map(manager_tb_mapping).fillna("")
            else:
                assignments["ТБ"] = ""
        
        log_debug(
            logger,
            f"{scenario_name}: назначено {len(assignments)} записей (режим latest)",
            class_name="Scenario",
            func_name="build_assignment_table",
        )
        return assignments

    default_name = defaults["manager_name"]
    default_id = _ensure_manager_identity(
        defaults["manager_id"], default_value=defaults["manager_id"], identifiers=identifiers
    )

    records: List[Dict[str, Any]] = []
    for row in variant_df.itertuples(index=False):
        base = {column: getattr(row, column, None) for column in key_columns}
        growth = getattr(row, "Прирост", 0.0) or 0.0
        fact_t0 = getattr(row, "Факт_T0", 0.0) or 0.0
        fact_t1 = getattr(row, "Факт_T1", 0.0) or 0.0

        manager_t0 = getattr(row, "Таб. номер ВКО_T0", None)
        manager_name_t0 = getattr(row, "ВКО_T0", None) or default_name
        manager_t0 = _ensure_manager_identity(
            manager_t0 or default_id, default_value=default_id, identifiers=identifiers
        )

        manager_t1 = getattr(row, "Таб. номер ВКО_T1", None)
        manager_name_t1 = getattr(row, "ВКО_T1", None) or default_name
        manager_t1 = _ensure_manager_identity(
            manager_t1 or default_id, default_value=default_id, identifiers=identifiers
        )

        if fact_t0 or growth > 0:
            record = {
                **base,
                SELECTED_MANAGER_ID_COL: manager_t0,
                SELECTED_MANAGER_NAME_COL: manager_name_t0,
                "Источник": "T0",
                "Факт_T0": fact_t0,
                "Факт_T1": 0.0,
                "Прирост": max(growth, 0.0),
            }
            # Добавляем ТБ, если его нет, определяя по табельному номеру
            if "ТБ" not in record and "tb" not in record:
                if manager_tb_mapping is not None:
                    record["ТБ"] = manager_tb_mapping.get(manager_t0, "")
                else:
                    record["ТБ"] = ""
            records.append(record)

        if fact_t1 or growth < 0:
            record = {
                **base,
                SELECTED_MANAGER_ID_COL: manager_t1,
                SELECTED_MANAGER_NAME_COL: manager_name_t1,
                "Источник": "T1",
                "Факт_T0": 0.0,
                "Факт_T1": fact_t1,
                "Прирост": min(growth, 0.0),
            }
            # Добавляем ТБ, если его нет, определяя по табельному номеру
            if "ТБ" not in record and "tb" not in record:
                if manager_tb_mapping is not None:
                    record["ТБ"] = manager_tb_mapping.get(manager_t1, "")
                else:
                    record["ТБ"] = ""
            records.append(record)

    assignments = pd.DataFrame(records)
    if assignments.empty:
        columns = key_columns + [
            SELECTED_MANAGER_ID_COL,
            SELECTED_MANAGER_NAME_COL,
            "Источник",
            "Факт_T0",
            "Факт_T1",
            "Прирост",
        ]
        assignments = pd.DataFrame(columns=columns)

    log_debug(
        logger,
        f"{scenario_name}: назначено {len(assignments)} записей (режим per_file)",
        class_name="Scenario",
        func_name="build_assignment_table",
    )
    return assignments


def build_manager_tb_mapping(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
) -> pd.Series:
    """Строит словарь соответствия табельного номера менеджера и ТБ из исходных данных."""
    
    # Собираем данные из обоих датафреймов
    dataframes_to_concat = []
    
    # Добавляем данные из current_df, если он не пустой и содержит нужные колонки
    if not current_df.empty and "manager_id" in current_df.columns and "tb" in current_df.columns:
        dataframes_to_concat.append(current_df[["manager_id", "tb"]].drop_duplicates())
    
    # Добавляем данные из previous_df, если он не пустой и содержит нужные колонки
    if not previous_df.empty and "manager_id" in previous_df.columns and "tb" in previous_df.columns:
        dataframes_to_concat.append(previous_df[["manager_id", "tb"]].drop_duplicates())
    
    # Объединяем все датафреймы
    if dataframes_to_concat:
        combined = pd.concat(dataframes_to_concat, ignore_index=True).drop_duplicates()
        # Если у одного менеджера несколько ТБ, берём первое (можно изменить логику на most_common)
        mapping = combined.groupby("manager_id")["tb"].first()
    else:
        # Если нет данных, возвращаем пустой Series
        mapping = pd.Series(dtype=object, name="tb")
    
    return mapping


def build_manager_gosb_mapping(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
) -> pd.Series:
    """Строит словарь соответствия табельного номера менеджера и ГОСБ из исходных данных."""
    
    # Собираем данные из обоих датафреймов
    dataframes_to_concat = []
    
    # Добавляем данные из current_df, если он не пустой и содержит нужные колонки
    if not current_df.empty and "manager_id" in current_df.columns and "gosb" in current_df.columns:
        dataframes_to_concat.append(current_df[["manager_id", "gosb"]].drop_duplicates())
    
    # Добавляем данные из previous_df, если он не пустой и содержит нужные колонки
    if not previous_df.empty and "manager_id" in previous_df.columns and "gosb" in previous_df.columns:
        dataframes_to_concat.append(previous_df[["manager_id", "gosb"]].drop_duplicates())
    
    # Объединяем все датафреймы
    if dataframes_to_concat:
        combined = pd.concat(dataframes_to_concat, ignore_index=True).drop_duplicates()
        # Если у одного менеджера несколько ГОСБ, берём первое
        mapping = combined.groupby("manager_id")["gosb"].first()
    else:
        # Если нет данных, возвращаем пустой Series
        mapping = pd.Series(dtype=object, name="gosb")
    
    return mapping


def build_client_summary_by_inn(
    variant_df: pd.DataFrame,
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    previous2_df: Optional[pd.DataFrame],
    manager_tb_mapping: pd.Series,
    manager_gosb_mapping: pd.Series,
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Строит свод по ИНН с детальной информацией по клиентам.
    
    Для каждого клиента показывает:
    - ИНН
    - Кол-во разных ТН в T-0, T-1, T-2
    - Какой ТН был выбран в каждом файле
    - Итоговый ТН с ФИО и ТБ
    
    Args:
        variant_df: DataFrame с данными варианта (до агрегации)
        current_df: DataFrame с данными T-0
        previous_df: DataFrame с данными T-1
        previous2_df: DataFrame с данными T-2 (может быть None)
        manager_tb_mapping: Маппинг табельного номера на ТБ
        manager_gosb_mapping: Маппинг табельного номера на ГОСБ
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с колонками: ИНН, Кол-во ТН_T0, ТН_T0, Кол-во ТН_T1, ТН_T1, 
        Кол-во ТН_T2 (если есть), ТН_T2 (если есть), 
        Итоговый ТН, ФИО КМ, ТБ
    """
    log_debug(
        logger,
        "Строю свод по ИНН",
        class_name="Aggregator",
        func_name="build_client_summary_by_inn",
    )
    
    # Определяем колонку клиента
    client_col = "client_id" if "client_id" in variant_df.columns else variant_df.columns[0]
    
    # Подсчитываем количество разных ТН для каждого клиента в каждом файле
    def count_unique_managers(df: pd.DataFrame, client_col: str, suffix: str) -> pd.Series:
        """Подсчитывает количество уникальных ТН для каждого клиента."""
        if df.empty:
            return pd.Series(dtype=int, name=f"Кол-во ТН_{suffix}")
        return df.groupby(client_col)["manager_id"].nunique()
    
    # Подсчитываем сумму фактов для каждого клиента в каждом файле
    def sum_facts(df: pd.DataFrame, client_col: str, suffix: str) -> pd.Series:
        """Подсчитывает сумму фактов для каждого клиента."""
        if df.empty or "fact_value_clean" not in df.columns:
            return pd.Series(dtype=float, name=f"Факт_{suffix}")
        return df.groupby(client_col)["fact_value_clean"].sum()
    
    # Подсчитываем для каждого файла
    count_t0_series = count_unique_managers(current_df, client_col, "T0")
    count_t1_series = count_unique_managers(previous_df, client_col, "T1")
    
    # Подсчитываем сумму фактов для каждого файла
    fact_t0_series = sum_facts(current_df, client_col, "T0")
    fact_t1_series = sum_facts(previous_df, client_col, "T1")
    
    # Начинаем с variant_df и добавляем колонки
    result = variant_df[[client_col]].drop_duplicates().copy()
    
    # Добавляем количество ТН
    result = result.set_index(client_col)
    result["Кол-во ТН_T0"] = count_t0_series
    result["Кол-во ТН_T1"] = count_t1_series
    if previous2_df is not None and not previous2_df.empty:
        count_t2_series = count_unique_managers(previous2_df, client_col, "T2")
        result["Кол-во ТН_T2"] = count_t2_series
        fact_t2_series = sum_facts(previous2_df, client_col, "T2")
        result["Факт_T2"] = fact_t2_series
    
    # Добавляем факты
    result["Факт_T0"] = fact_t0_series
    result["Факт_T1"] = fact_t1_series
    
    # ВАЖНО: заполняем NaN нулями ПЕРЕД расчетом прироста
    # Если клиента нет в каком-то периоде, факт должен быть 0, а не NaN
    result["Факт_T0"] = result["Факт_T0"].fillna(0.0)
    result["Факт_T1"] = result["Факт_T1"].fillna(0.0)
    if "Факт_T2" in result.columns:
        result["Факт_T2"] = result["Факт_T2"].fillna(0.0)
    
    # Вычисляем прирост
    # Формула с T-2: прирост = (T-0 - T-1) - (T-1 - T-2) = T0 - 2*T1 + T2
    # Если есть только T-0 (T-1=0, T-2=0), то прирост = T0 - 2*0 + 0 = T0
    if previous2_df is not None and not previous2_df.empty and "Факт_T2" in result.columns:
        result["Прирост"] = (result["Факт_T0"] - result["Факт_T1"]) - (result["Факт_T1"] - result["Факт_T2"])
    else:
        result["Прирост"] = result["Факт_T0"] - result["Факт_T1"]
    
    result = result.reset_index()
    
    # Добавляем выбранные ТН из variant_df (без ФИО и ТБ для промежуточных)
    if "Таб. номер ВКО_T0" in variant_df.columns:
        tn_t0 = variant_df[[client_col, "Таб. номер ВКО_T0"]].drop_duplicates()
        result = result.merge(tn_t0, on=client_col, how="left")
        result = result.rename(columns={"Таб. номер ВКО_T0": "ТН_T0"})
    
    if "Таб. номер ВКО_T1" in variant_df.columns:
        tn_t1 = variant_df[[client_col, "Таб. номер ВКО_T1"]].drop_duplicates()
        result = result.merge(tn_t1, on=client_col, how="left")
        result = result.rename(columns={"Таб. номер ВКО_T1": "ТН_T1"})
    
    if previous2_df is not None and "Таб. номер ВКО_T2" in variant_df.columns:
        tn_t2 = variant_df[[client_col, "Таб. номер ВКО_T2"]].drop_duplicates()
        result = result.merge(tn_t2, on=client_col, how="left")
        result = result.rename(columns={"Таб. номер ВКО_T2": "ТН_T2"})
    
    # Добавляем итоговый ТН с ФИО и ТБ
    if "Таб. номер ВКО_Актуальный" in variant_df.columns:
        final_tn = variant_df[[client_col, "Таб. номер ВКО_Актуальный", "ВКО_Актуальный"]].drop_duplicates()
        result = result.merge(final_tn, on=client_col, how="left")
        
        # Добавляем ТБ для итогового ТН
        result["ТБ"] = result["Таб. номер ВКО_Актуальный"].map(manager_tb_mapping).fillna("")
    
    # Переименовываем колонки для читаемости
    rename_map = {
        client_col: "ИНН",
        "Таб. номер ВКО_Актуальный": "Итоговый ТН",
        "ВКО_Актуальный": "ФИО КМ",
    }
    result = result.rename(columns=rename_map)
    
    # Заполняем пропуски в числовых колонках
    for col in ["Кол-во ТН_T0", "Кол-во ТН_T1"]:
        if col in result.columns:
            result[col] = result[col].fillna(0).astype(int)
    if "Кол-во ТН_T2" in result.columns:
        result["Кол-во ТН_T2"] = result["Кол-во ТН_T2"].fillna(0).astype(int)
    
    # Заполняем пропуски в числовых колонках фактов и прироста
    for col in ["Факт_T0", "Факт_T1", "Прирост"]:
        if col in result.columns:
            result[col] = result[col].fillna(0.0).astype(float)
    if "Факт_T2" in result.columns:
        result["Факт_T2"] = result["Факт_T2"].fillna(0.0).astype(float)
    
    # Переупорядочиваем колонки
    base_cols = ["ИНН"]
    if "Кол-во ТН_T0" in result.columns:
        base_cols.extend(["Кол-во ТН_T0", "ТН_T0", "Факт_T0"])
    if "Кол-во ТН_T1" in result.columns:
        base_cols.extend(["Кол-во ТН_T1", "ТН_T1", "Факт_T1"])
    if "Кол-во ТН_T2" in result.columns:
        base_cols.extend(["Кол-во ТН_T2", "ТН_T2", "Факт_T2"])
    base_cols.extend(["Прирост", "Итоговый ТН", "ФИО КМ", "ТБ"])
    
    # Оставляем только существующие колонки
    existing_cols = [col for col in base_cols if col in result.columns]
    other_cols = [col for col in result.columns if col not in existing_cols]
    result = result[existing_cols + other_cols]
    
    log_debug(
        logger,
        f"Свод по ИНН: подготовлено {len(result)} строк",
        class_name="Aggregator",
        func_name="build_client_summary_by_inn",
    )
    
    return result


def build_assignment_summary(
    assignment_df: pd.DataFrame,
    *,
    include_tb: bool,
    logger: Mapping[str, Any],
    summary_name: str,
    manager_tb_mapping: Optional[pd.Series] = None,
) -> pd.DataFrame:
    """Суммирует факты/приросты по выбранным менеджерам. Всегда добавляет ТБ, определяя его по табельному номеру."""

    group_columns = [SELECTED_MANAGER_ID_COL, SELECTED_MANAGER_NAME_COL]
    
    # Всегда добавляем ТБ, определяя его по табельному номеру, если его нет в данных
    if "ТБ" not in assignment_df.columns and "tb" not in assignment_df.columns:
        if manager_tb_mapping is not None and not assignment_df.empty:
            # Определяем ТБ по табельному номеру
            assignment_df = assignment_df.copy()
            assignment_df["ТБ"] = assignment_df[SELECTED_MANAGER_ID_COL].map(manager_tb_mapping).fillna("")
    
    tb_column_name: Optional[str] = None
    if "ТБ" in assignment_df.columns:
        tb_column_name = "ТБ"
    elif "tb" in assignment_df.columns:
        tb_column_name = "tb"
    
    if tb_column_name:
        group_columns.append(tb_column_name)

    if assignment_df.empty:
        columns = group_columns + ["Факт_T0", "Факт_T1", "Прирост", "Количество записей"]
        return pd.DataFrame(columns=columns)

    numeric_columns = ["Факт_T0", "Факт_T1", "Прирост"]
    summary = (
        assignment_df.groupby(group_columns, dropna=False)[numeric_columns]
        .sum()
        .reset_index()
    )
    counts = assignment_df.groupby(group_columns, dropna=False).size().reset_index(name="Количество записей")
    summary = summary.merge(counts, on=group_columns, how="left")
    if tb_column_name == "tb":
        summary = summary.rename(columns={"tb": "ТБ"})

    log_debug(
        logger,
        f"{summary_name}: сформирована сводная таблица на {len(summary)} менеджеров",
        class_name="Scenario",
        func_name="build_assignment_summary",
    )
    return summary


def normalize_string(value: Any) -> str:
    """Возвращает очищенную строку без None."""

    if value is None:
        return ""
    return str(value).strip()


# ==================== КЛАССЫ ООП ====================

class DataLoader:
    """Класс для загрузки и очистки данных из Excel файлов.
    
    Инкапсулирует логику чтения файлов, переименования колонок, форматирования
    идентификаторов и фильтрации запрещенных значений.
    
    Атрибуты:
        identifiers: Настройки форматирования идентификаторов (manager_id, client_id)
        logger: Логгер для записи сообщений
    """
    
    def __init__(self, identifiers: Mapping[str, Mapping[str, Any]], logger: Mapping[str, Any]):
        """Инициализирует загрузчик данных.
        
        Args:
            identifiers: Словарь с настройками форматирования идентификаторов
            logger: Логгер с методами info и debug
        """
        self.identifiers = identifiers
        self.logger = logger
    
    def read_source_file(
        self,
        file_path: Path,
        sheet_name: str,
        columns: List[Dict[str, str]],
        drop_rules: Mapping[str, Iterable[str]],
    ) -> pd.DataFrame:
        """Загружает исходный Excel и подготавливает данные.
        
        Выполняет следующие операции:
        1. Читает указанный лист Excel
        2. Переименовывает колонки согласно маппингу
        3. Нормализует строковые поля
        4. Форматирует идентификаторы (табельные номера и ИНН)
        5. Преобразует факт в числовой формат
        6. Удаляет строки с запрещенными значениями
        
        Args:
            file_path: Путь к файлу Excel
            sheet_name: Имя листа для чтения
            columns: Список словарей с alias и source для колонок
            drop_rules: Правила фильтрации строк (словарь {alias: tuple(values)})
        
        Returns:
            DataFrame с очищенными и отформатированными данными
        
        Raises:
            FileNotFoundError: Если файл не существует
            ValueError: Если отсутствуют обязательные колонки
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {file_path}")

        log_info(self.logger, f"Загружаю данные из файла {file_path.name}")
        
        # Формируем маппинг колонок из списка
        column_maps = {column["source"]: column["alias"] for column in columns}
        
        # Читаем один лист Excel и сразу переименовываем колонки в единый формат
        raw_df = pd.read_excel(file_path, sheet_name=sheet_name)
        renamed = raw_df.rename(columns=column_maps)

        required_columns = list(column_maps.values())
        missing = [col for col in required_columns if col not in renamed.columns]
        if missing:
            raise ValueError(
                f"Отсутствуют обязательные колонки {missing} в файле {file_path}"
            )

        prepared = renamed[required_columns].copy()

        # Строковые столбцы очищаем от пробелов и None
        for column in ("tb", "gosb", "manager_name"):
            prepared[column] = prepared[column].apply(normalize_string)

        manager_identifier = self.identifiers["manager_id"]
        client_identifier = self.identifiers["client_id"]

        # Форматируем табельные номера и ИНН в заранее заданную длину
        prepared["manager_id"] = prepared["manager_id"].apply(
            lambda value: format_identifier(
                value=value,
                total_length=manager_identifier["total_length"],
                fill_char=manager_identifier["fill_char"],
            )
        )
        prepared["client_id"] = prepared["client_id"].apply(
            lambda value: format_identifier(
                value=value,
                total_length=client_identifier["total_length"],
                fill_char=client_identifier["fill_char"],
            )
        )

        prepared["fact_value_clean"] = prepared["fact_value"].apply(safe_to_float)

        cleaned = self.drop_forbidden_rows(prepared, drop_rules)
        log_debug(
            self.logger,
            f"После очистки в {file_path.name} осталось строк: {len(cleaned)}",
            class_name="DataLoader",
            func_name="read_source_file",
        )
        return cleaned
    
    def apply_in_rules(
        self,
        df: pd.DataFrame,
        in_rules: List[Dict[str, Any]],
    ) -> pd.DataFrame:
        """Применяет правила включения строк (IN фильтры).
        
        Строка попадает в результат только если она проходит ВСЕ условия из in_rules (И).
        Условия применяются после drop_rules (исключается из DROP).
        
        Args:
            df: DataFrame для фильтрации
            in_rules: Список правил включения [{"alias": "column", "values": [...], "condition": "in" или "not_in"}]
        
        Returns:
            DataFrame с отфильтрованными строками
        """
        if not in_rules:
            return df
        
        filtered = df.copy()
        
        for rule in in_rules:
            column = rule.get("alias")
            if column not in filtered.columns:
                log_debug(
                    self.logger,
                    f"Колонка {column} отсутствует в данных, пропускаем IN правило",
                    class_name="DataLoader",
                    func_name="apply_in_rules",
                )
                continue
            
            values = rule.get("values", [])
            condition = rule.get("condition", "in")
            
            if not values:
                continue
            
            # Нормализуем значения для сравнения
            normalized_values = {str(v).strip().lower() for v in values}
            
            if condition == "in":
                # Значение должно быть в списке
                mask = filtered[column].apply(
                    lambda x: str(x).strip().lower() in normalized_values if pd.notna(x) else False
                )
            elif condition == "not_in":
                # Значение НЕ должно быть в списке
                mask = filtered[column].apply(
                    lambda x: str(x).strip().lower() not in normalized_values if pd.notna(x) else True
                )
            else:
                log_debug(
                    self.logger,
                    f"Неизвестное условие '{condition}' в IN правиле для колонки {column}, пропускаем",
                    class_name="DataLoader",
                    func_name="apply_in_rules",
                )
                continue
            
            before = len(filtered)
            filtered = filtered[mask]
            log_debug(
                self.logger,
                f"Колонка {column} (IN правило, condition={condition}): оставлено {len(filtered)} из {before} строк",
                class_name="DataLoader",
                func_name="apply_in_rules",
            )
        
        return filtered
    
    def drop_forbidden_rows(
        self,
        df: pd.DataFrame,
        drop_rules: Mapping[str, Dict[str, Any]],
    ) -> pd.DataFrame:
        """Удаляет строки с запрещёнными значениями с поддержкой условной логики.
        
        Проходит по каждой колонке из drop_rules и удаляет строки, где значение
        (приведенное к нижнему регистру) совпадает с одним из запрещенных.
        
        Поддерживает условное удаление:
        - remove_unconditionally: убирать всегда (по умолчанию True)
        - check_by_inn: если True, не убираем строку, если по этому ИНН есть строки с другими значениями
        - check_by_tn: если True, не убираем строку, если по этому ТН есть строки с другими значениями
        - Если оба check_by_inn и check_by_tn True - работает как ИЛИ
        
        Args:
            df: DataFrame для очистки
            drop_rules: Словарь {column_alias: {values: tuple, remove_unconditionally: bool, check_by_inn: bool, check_by_tn: bool}}
        
        Returns:
            DataFrame без запрещенных строк
        """
        cleaned = df.copy()
        
        for column, rule in drop_rules.items():
            if column not in cleaned.columns:
                log_debug(
                    self.logger,
                    f"Колонка {column} отсутствует в данных, пропускаем правило",
                    class_name="DataLoader",
                    func_name="drop_forbidden_rows",
                )
                continue
            
            values = rule.get("values", ())
            remove_unconditionally = rule.get("remove_unconditionally", True)
            check_by_inn = rule.get("check_by_inn", False)
            check_by_tn = rule.get("check_by_tn", False)
            
            forbidden = {value.lower() for value in values}

            def is_forbidden_value(value: Any) -> bool:
                """Проверяет, является ли значение запрещенным."""
                if value is None:
                    return False
                return str(value).strip().lower() in forbidden

            # Находим строки с запрещенными значениями
            mask_forbidden = cleaned[column].apply(is_forbidden_value)
            
            if not mask_forbidden.any():
                log_debug(
                    self.logger,
                    f"Колонка {column}: запрещенных значений не найдено",
                    class_name="DataLoader",
                    func_name="drop_forbidden_rows",
                )
                continue
            
            if not remove_unconditionally:
                # Если remove_unconditionally=False, не удаляем строки
                log_debug(
                    self.logger,
                    f"Колонка {column}: remove_unconditionally=False, строки не удаляются",
                    class_name="DataLoader",
                    func_name="drop_forbidden_rows",
                )
                continue
            
            if not check_by_inn and not check_by_tn:
                # Простое удаление без условий (старая логика)
                before = len(cleaned)
                cleaned = cleaned[~mask_forbidden]
                log_debug(
                    self.logger,
                    f"Колонка {column}: удалено {before - len(cleaned)} строк (безусловно)",
                    class_name="DataLoader",
                    func_name="drop_forbidden_rows",
                )
            else:
                # Условное удаление: удаляем строки, но если условие выполняется (есть другие значения по ИНН/ТН), не удаляем
                rows_to_remove = mask_forbidden.copy()
                
                # Проверяем условия для каждой строки с запрещенным значением
                for idx in cleaned[mask_forbidden].index:
                    row = cleaned.loc[idx]
                    should_keep = False
                    
                    if check_by_inn and "client_id" in cleaned.columns:
                        # Проверяем, есть ли по этому ИНН строки с другими значениями в этой колонке
                        client_id = row["client_id"]
                        if pd.notna(client_id):
                            other_rows_same_inn = cleaned[
                                (cleaned["client_id"] == client_id) & 
                                (cleaned.index != idx)
                            ]
                            if len(other_rows_same_inn) > 0:
                                # Проверяем, есть ли среди них строки с другими значениями в этой колонке
                                other_values = other_rows_same_inn[column].apply(
                                    lambda v: not is_forbidden_value(v) if pd.notna(v) else False
                                )
                                if other_values.any():
                                    should_keep = True
                    
                    if check_by_tn and "manager_id" in cleaned.columns:
                        # Проверяем, есть ли по этому ТН строки с другими значениями в этой колонке
                        manager_id = row["manager_id"]
                        if pd.notna(manager_id):
                            other_rows_same_tn = cleaned[
                                (cleaned["manager_id"] == manager_id) & 
                                (cleaned.index != idx)
                            ]
                            if len(other_rows_same_tn) > 0:
                                # Проверяем, есть ли среди них строки с другими значениями в этой колонке
                                other_values = other_rows_same_tn[column].apply(
                                    lambda v: not is_forbidden_value(v) if pd.notna(v) else False
                                )
                                if other_values.any():
                                    should_keep = True
                    
                    # Если хотя бы одно условие выполняется (ИЛИ), не убираем строку
                    if should_keep:
                        rows_to_remove.loc[idx] = False
                
                before = len(cleaned)
                cleaned = cleaned[~rows_to_remove]
                log_debug(
                    self.logger,
                    f"Колонка {column}: удалено {before - len(cleaned)} строк "
                    f"(условно: remove_unconditionally={remove_unconditionally}, "
                    f"check_by_inn={check_by_inn}, check_by_tn={check_by_tn})",
                    class_name="DataLoader",
                    func_name="drop_forbidden_rows",
                )
        
        return cleaned


class Aggregator:
    """Класс для агрегации данных и определения менеджеров.
    
    Инкапсулирует логику группировки данных, суммирования фактов,
    определения доминантных менеджеров и построения наборов данных для вариантов.
    
    Атрибуты:
        defaults: Настройки по умолчанию (manager_name, manager_id)
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    """
    
    def __init__(
        self,
        defaults: Mapping[str, Any],
        identifiers: Mapping[str, Any],
        logger: Mapping[str, Any],
    ):
        """Инициализирует агрегатор.
        
        Args:
            defaults: Словарь с настройками по умолчанию
            identifiers: Словарь с настройками форматирования идентификаторов
            logger: Логгер с методами info и debug
        """
        self.defaults = defaults
        self.identifiers = identifiers
        self.logger = logger
    
    def aggregate_facts(
        self,
        df: pd.DataFrame,
        key_columns: List[str],
        suffix: str,
        variant_name: str,
    ) -> pd.DataFrame:
        """Группирует данные по ключу и суммирует факт.
        
        Группирует DataFrame по указанным ключевым колонкам и суммирует
        значения fact_value_clean. Результат переименовывается в Факт_{suffix}.
        
        Args:
            df: Исходный DataFrame с данными
            key_columns: Список колонок для группировки
            suffix: Суффикс для имени результирующей колонки (например, "T0", "T1")
            variant_name: Имя варианта для логирования
        
        Returns:
            DataFrame с колонками key_columns и Факт_{suffix}
        """
        grouped = (
            df[key_columns + ["fact_value_clean"]]
            .fillna({"fact_value_clean": 0.0})
            .groupby(key_columns, dropna=False, as_index=False)
            .sum(numeric_only=True)
        )
        renamed = grouped.rename(columns={"fact_value_clean": f"Факт_{suffix}"})
        log_debug(
            self.logger,
            f"{variant_name}: агрегировано {len(renamed)} строк для суффикса {suffix}",
            class_name="Aggregator",
            func_name="aggregate_facts",
        )
        return renamed
    
    def select_best_manager(
        self,
        df: pd.DataFrame,
        key_columns: List[str],
        variant_name: str,
    ) -> pd.DataFrame:
        """Определяет доминантного менеджера (по сумме факта) для каждого ключа.
        
        ВАЖНО: Возвращает менеджера даже если сумма факта = 0, если клиент есть в файле.
        Если клиента нет в файле (после фильтрации), то его не будет в результате.
        
        Алгоритм:
        1. Группирует данные по (ключ, manager_id, manager_name) и суммирует fact_value_clean
        2. Для каждого ключа выбирает менеджера с максимальной суммой
        3. Если суммы равны (включая случай, когда все суммы = 0), pandas idxmax вернёт первую попавшуюся
        4. Если клиента нет в файле (после фильтрации), то его не будет в результате
        
        Args:
            df: Исходный DataFrame с данными (уже отфильтрованный)
            key_columns: Список колонок для ключа (например, ["client_id"] или ["client_id", "tb"])
            variant_name: Имя варианта для логирования
        
        Returns:
            DataFrame с колонками key_columns, "ВКО", "Таб. номер ВКО"
            Содержит только те ключи, которые есть в df (после фильтрации).
            Если клиента нет в df, его не будет в результате.
        """
        # Если DataFrame пустой, возвращаем пустой результат
        if df.empty:
            return pd.DataFrame(columns=key_columns + ["ВКО", "Таб. номер ВКО"])
        
        additional_columns = [
            column for column in ("manager_name", "manager_id") if column not in key_columns
        ]
        grouping_columns = key_columns + additional_columns
        grouped = (
            df[grouping_columns + ["fact_value_clean"]]
            .fillna({"fact_value_clean": 0.0})
            .groupby(grouping_columns, dropna=False, as_index=False)
            .sum(numeric_only=True)
        )
        # idxmax вернет индекс даже если все значения = 0 (если клиент есть в файле)
        idx = grouped.groupby(key_columns, dropna=False)["fact_value_clean"].idxmax()
        best = grouped.loc[idx, key_columns + additional_columns].copy()
        result = best.copy()
        if "manager_name" in result.columns and "manager_name" not in key_columns:
            result = result.rename(columns={"manager_name": "ВКО"})
        if "manager_id" in key_columns and "manager_id" in result.columns:
            result["Таб. номер ВКО"] = result["manager_id"]
        elif "manager_id" in result.columns:
            result = result.rename(columns={"manager_id": "Таб. номер ВКО"})
        log_debug(
            self.logger,
            f"{variant_name}: выбраны менеджеры для {len(result)} ключей",
            class_name="Aggregator",
            func_name="select_best_manager",
        )
        return result
    
    def build_latest_manager(
        self,
        current_best: pd.DataFrame,
        previous_best: pd.DataFrame,
        key_columns: List[str],
        variant_name: str,
    ) -> pd.DataFrame:
        """Комбинирует менеджеров, отдавая приоритет файлу T-0.
        
        Объединяет менеджеров из T-0 и T-1, приоритет отдается T-0.
        Если менеджер не найден, используется значение по умолчанию.
        
        Args:
            current_best: DataFrame с менеджерами из T-0 (колонки: key_columns, "ВКО", "Таб. номер ВКО")
            previous_best: DataFrame с менеджерами из T-1 (колонки: key_columns, "ВКО", "Таб. номер ВКО")
            key_columns: Список колонок для ключа
            variant_name: Имя варианта для логирования
        
        Returns:
            DataFrame с колонками key_columns, "ВКО_Актуальный", "Таб. номер ВКО_Актуальный"
        """
        default_name = self.defaults["manager_name"]
        identifier_settings = self.identifiers["manager_id"]
        default_id = format_identifier(
            self.defaults["manager_id"],
            total_length=identifier_settings["total_length"],
            fill_char=identifier_settings["fill_char"],
        )

        curr = (
            current_best.set_index(key_columns)
            if not current_best.empty
            else pd.DataFrame(columns=key_columns + ["ВКО", "Таб. номер ВКО"]).set_index(key_columns)
        )
        prev = (
            previous_best.set_index(key_columns)
            if not previous_best.empty
            else pd.DataFrame(columns=key_columns + ["ВКО", "Таб. номер ВКО"]).set_index(key_columns)
        )

        # ВАЖНО: переименовываем колонки ПЕРЕД join, чтобы гарантировать правильные имена
        # Переименовываем колонки в каждом DataFrame (если они не пустые)
        if not prev.empty:
            prev_renamed = prev.rename(columns={"ВКО": "ВКО_prev", "Таб. номер ВКО": "Таб. номер ВКО_prev"})
        else:
            prev_renamed = pd.DataFrame(columns=key_columns + ["ВКО_prev", "Таб. номер ВКО_prev"]).set_index(key_columns)
        
        if not curr.empty:
            curr_renamed = curr.rename(columns={"ВКО": "ВКО_curr", "Таб. номер ВКО": "Таб. номер ВКО_curr"})
        else:
            curr_renamed = pd.DataFrame(columns=key_columns + ["ВКО_curr", "Таб. номер ВКО_curr"]).set_index(key_columns)
        
        # Теперь объединяем с правильными именами колонок
        combined = prev_renamed.join(curr_renamed, how="outer")
        
        # Проверяем наличие колонок перед обращением (теперь имена колонок гарантированы)
        vko_curr = combined.get("ВКО_curr", pd.Series(index=combined.index, dtype=object))
        vko_prev = combined.get("ВКО_prev", pd.Series(index=combined.index, dtype=object))
        combined["ВКО_Актуальный"] = vko_curr.combine_first(vko_prev).fillna(default_name)
        
        tab_curr = combined.get("Таб. номер ВКО_curr", pd.Series(index=combined.index, dtype=object))
        tab_prev = combined.get("Таб. номер ВКО_prev", pd.Series(index=combined.index, dtype=object))
        combined["Таб. номер ВКО_Актуальный"] = tab_curr.combine_first(tab_prev).fillna(default_id)

        result = combined.reset_index()[key_columns + ["ВКО_Актуальный", "Таб. номер ВКО_Актуальный"]]
        log_debug(
            self.logger,
            f"{variant_name}: определены актуальные менеджеры для {len(result)} ключей",
            class_name="Aggregator",
            func_name="build_latest_manager",
        )
        return result
    
    def build_latest_manager_with_t2(
        self,
        current_best: pd.DataFrame,
        previous_best: pd.DataFrame,
        previous2_best: Optional[pd.DataFrame],
        key_columns: List[str],
        variant_name: str,
    ) -> pd.DataFrame:
        """Комбинирует менеджеров, отдавая приоритет файлу T-0, затем T-1, затем T-2.
        
        Объединяет менеджеров из всех трех периодов, приоритет: T-0 → T-1 → T-2.
        Если менеджер не найден, используется значение по умолчанию.
        
        Args:
            current_best: DataFrame с менеджерами из T-0
            previous_best: DataFrame с менеджерами из T-1
            previous2_best: DataFrame с менеджерами из T-2 (может быть None)
            key_columns: Список колонок для ключа
            variant_name: Имя варианта для логирования
        
        Returns:
            DataFrame с колонками key_columns, "ВКО_Актуальный", "Таб. номер ВКО_Актуальный"
        """
        default_name = self.defaults["manager_name"]
        identifier_settings = self.identifiers["manager_id"]
        default_id = format_identifier(
            self.defaults["manager_id"],
            total_length=identifier_settings["total_length"],
            fill_char=identifier_settings["fill_char"],
        )

        curr = (
            current_best.set_index(key_columns)
            if not current_best.empty
            else pd.DataFrame(columns=key_columns + ["ВКО", "Таб. номер ВКО"]).set_index(key_columns)
        )
        prev = (
            previous_best.set_index(key_columns)
            if not previous_best.empty
            else pd.DataFrame(columns=key_columns + ["ВКО", "Таб. номер ВКО"]).set_index(key_columns)
        )
        prev2 = (
            previous2_best.set_index(key_columns)
            if previous2_best is not None and not previous2_best.empty
            else pd.DataFrame(columns=key_columns + ["ВКО", "Таб. номер ВКО"]).set_index(key_columns)
        )

        # Объединяем все ключи из всех трех файлов (outer join)
        # Приоритет: T-0 (curr) → T-1 (prev) → T-2 (prev2)
        # ВАЖНО: переименовываем колонки ПЕРЕД join, чтобы гарантировать правильные имена
        # Сначала переименовываем колонки в каждом DataFrame (если они не пустые)
        if not prev2.empty:
            prev2_renamed = prev2.rename(columns={"ВКО": "ВКО_prev2", "Таб. номер ВКО": "Таб. номер ВКО_prev2"})
        else:
            prev2_renamed = pd.DataFrame(columns=key_columns + ["ВКО_prev2", "Таб. номер ВКО_prev2"]).set_index(key_columns)
        
        if not prev.empty:
            prev_renamed = prev.rename(columns={"ВКО": "ВКО_prev", "Таб. номер ВКО": "Таб. номер ВКО_prev"})
        else:
            prev_renamed = pd.DataFrame(columns=key_columns + ["ВКО_prev", "Таб. номер ВКО_prev"]).set_index(key_columns)
        
        if not curr.empty:
            curr_renamed = curr.rename(columns={"ВКО": "ВКО_curr", "Таб. номер ВКО": "Таб. номер ВКО_curr"})
        else:
            curr_renamed = pd.DataFrame(columns=key_columns + ["ВКО_curr", "Таб. номер ВКО_curr"]).set_index(key_columns)
        
        # Теперь объединяем с правильными именами колонок
        combined = prev2_renamed.join(prev_renamed, how="outer")
        combined = combined.join(curr_renamed, how="outer")
        
        # Определяем актуального менеджера: приоритет curr (T-0) → prev (T-1) → prev2 (T-2)
        # combine_first берет значение из первой серии, если оно не NaN, иначе из второй, и т.д.
        # Проверяем наличие колонок перед обращением
        vko_curr = combined.get("ВКО_curr", pd.Series(index=combined.index, dtype=object))
        vko_prev = combined.get("ВКО_prev", pd.Series(index=combined.index, dtype=object))
        vko_prev2 = combined.get("ВКО_prev2", pd.Series(index=combined.index, dtype=object))
        
        # Приоритет: сначала curr (T-0), затем prev (T-1), затем prev2 (T-2)
        # Если в T-0 есть значение - берем его, иначе смотрим T-1, иначе T-2
        combined["ВКО_Актуальный"] = (
            vko_curr
            .combine_first(vko_prev)
            .combine_first(vko_prev2)
            .fillna(default_name)
        )
        
        tab_curr = combined.get("Таб. номер ВКО_curr", pd.Series(index=combined.index, dtype=object))
        tab_prev = combined.get("Таб. номер ВКО_prev", pd.Series(index=combined.index, dtype=object))
        tab_prev2 = combined.get("Таб. номер ВКО_prev2", pd.Series(index=combined.index, dtype=object))
        
        # Приоритет: сначала curr (T-0), затем prev (T-1), затем prev2 (T-2)
        combined["Таб. номер ВКО_Актуальный"] = (
            tab_curr
            .combine_first(tab_prev)
            .combine_first(tab_prev2)
            .fillna(default_id)
        )

        result = combined.reset_index()[key_columns + ["ВКО_Актуальный", "Таб. номер ВКО_Актуальный"]]
        log_debug(
            self.logger,
            f"{variant_name}: определены актуальные менеджеры для {len(result)} ключей (T-0 → T-1 → T-2)",
            class_name="Aggregator",
            func_name="build_latest_manager_with_t2",
        )
        return result
    
    def assemble_variant_dataset_with_t2(
        self,
        variant_name: str,
        key_columns: List[str],
        current_df: pd.DataFrame,
        previous_df: pd.DataFrame,
        previous2_df: Optional[pd.DataFrame],
    ) -> pd.DataFrame:
        """Формирует таблицу для конкретного варианта ключа с поддержкой T-2.
        
        Выполняет следующие операции:
        1. Агрегирует факты по ключу для каждого периода (T-0, T-1, T-2)
        2. Объединяет результаты и рассчитывает прирост:
           - Если T-2 указан: прирост = (T-0 - T-1) - (T-1 - T-2)
           - Иначе: прирост = T-0 - T-1
        3. Определяет лучшего менеджера для каждого периода
        4. Определяет актуального менеджера (приоритет T-0 → T-1 → T-2)
        
        Args:
            variant_name: Имя варианта для логирования
            key_columns: Список колонок для ключа агрегации
            current_df: DataFrame с данными T-0
            previous_df: DataFrame с данными T-1
            previous2_df: DataFrame с данными T-2 (может быть None)
        
        Returns:
            DataFrame с колонками key_columns, Факт_T0, Факт_T1, Факт_T2 (если есть),
            Прирост, ВКО_T0, ВКО_T1, ВКО_T2 (если есть), ВКО_Актуальный, Таб. номер ВКО_Актуальный
        """
        log_debug(
            self.logger,
            f"{variant_name}: старт построения набора данных (T-2: {'да' if previous2_df is not None else 'нет'})",
            class_name="Aggregator",
            func_name="assemble_variant_dataset_with_t2",
        )

        agg_current = self.aggregate_facts(current_df, key_columns, "T0", variant_name)
        agg_previous = self.aggregate_facts(previous_df, key_columns, "T1", variant_name)
        
        if previous2_df is not None:
            agg_previous2 = self.aggregate_facts(previous2_df, key_columns, "T2", variant_name)
            # Объединяем все три периода
            merged = (
                pd.merge(agg_current, agg_previous, on=key_columns, how="outer")
                .merge(agg_previous2, on=key_columns, how="outer")
                .fillna({"Факт_T0": 0.0, "Факт_T1": 0.0, "Факт_T2": 0.0})
            )
            # Формула с T-2: прирост = (T-0 - T-1) - (T-1 - T-2)
            merged["Прирост"] = (merged["Факт_T0"] - merged["Факт_T1"]) - (merged["Факт_T1"] - merged["Факт_T2"])
        else:
            merged = (
                pd.merge(agg_current, agg_previous, on=key_columns, how="outer")
                .fillna({"Факт_T0": 0.0, "Факт_T1": 0.0})
            )
            merged["Прирост"] = merged["Факт_T0"] - merged["Факт_T1"]

        # Определяем лучшего менеджера для каждого периода
        best_current = self.select_best_manager(
            current_df, key_columns, variant_name
        ).rename(columns={"ВКО": "ВКО_T0", "Таб. номер ВКО": "Таб. номер ВКО_T0"})
        best_previous = self.select_best_manager(
            previous_df, key_columns, variant_name
        ).rename(columns={"ВКО": "ВКО_T1", "Таб. номер ВКО": "Таб. номер ВКО_T1"})
        
        merged = merged.merge(best_current, on=key_columns, how="left")
        merged = merged.merge(best_previous, on=key_columns, how="left")
        
        if previous2_df is not None:
            best_previous2 = self.select_best_manager(
                previous2_df, key_columns, variant_name
            ).rename(columns={"ВКО": "ВКО_T2", "Таб. номер ВКО": "Таб. номер ВКО_T2"})
            merged = merged.merge(best_previous2, on=key_columns, how="left")
            
            # Для определения актуального менеджера используем все ключи из merged
            # Приоритет: T-0 → T-1 → T-2
            # 
            # ЛОГИКА РАБОТЫ:
            # 1. select_best_manager возвращает менеджера для каждого ключа (ИНН), который есть в файле
            #    - Если клиент есть в файле (даже с суммой = 0), менеджер будет найден
            #    - Если клиента нет в файле (после фильтрации), его не будет в best_current/best_previous/best_previous2
            # 2. build_latest_manager_with_t2 делает outer join всех ключей из всех трех файлов
            #    - Если клиент есть только в T-0, он будет в best_current_renamed
            #    - combine_first берет значение из T-0 (приоритет T-0 → T-1 → T-2)
            # 3. Если ни в одном файле не найден менеджер, fillna заполнит значением по умолчанию
            # 
            # Сначала создаем DataFrame со всеми ключами из merged
            all_keys = merged[key_columns].drop_duplicates()
            
            # Переименовываем колонки для build_latest_manager_with_t2
            # ВАЖНО: НЕ делаем merge с all_keys перед вызовом build_latest_manager_with_t2,
            # потому что build_latest_manager_with_t2 сам делает outer join и объединяет все ключи
            best_current_renamed = best_current.rename(columns={"ВКО_T0": "ВКО", "Таб. номер ВКО_T0": "Таб. номер ВКО"})
            best_previous_renamed = best_previous.rename(columns={"ВКО_T1": "ВКО", "Таб. номер ВКО_T1": "Таб. номер ВКО"})
            best_previous2_renamed = best_previous2.rename(columns={"ВКО_T2": "ВКО", "Таб. номер ВКО_T2": "Таб. номер ВКО"})
            
            # Вызываем build_latest_manager_with_t2 с исходными данными
            # build_latest_manager_with_t2 сам делает outer join и объединяет все ключи из всех трех файлов
            # Если клиент есть только в T-0, он будет в best_current_renamed, и combine_first возьмет значение из T-0
            latest = self.build_latest_manager_with_t2(
                current_best=best_current_renamed,
                previous_best=best_previous_renamed,
                previous2_best=best_previous2_renamed,
                key_columns=key_columns,
                variant_name=variant_name,
            )
            
            # Убеждаемся, что latest содержит все ключи из merged
            # Если в latest нет ключа из merged, добавляем его с значениями по умолчанию
            latest = all_keys.merge(latest, on=key_columns, how="left")
            
            # Заполняем пропуски значениями по умолчанию ТОЛЬКО если менеджер не найден ни в одном файле
            # Если клиент есть только в T-0, то latest уже должен содержать правильное значение из T-0
            # после build_latest_manager_with_t2, поэтому fillna применяется только к действительно отсутствующим значениям
            default_name = self.defaults["manager_name"]
            identifier_settings = self.identifiers["manager_id"]
            default_id = format_identifier(
                self.defaults["manager_id"],
                total_length=identifier_settings["total_length"],
                fill_char=identifier_settings["fill_char"],
            )
            # Заполняем только если действительно нет значения (не должно быть после build_latest_manager_with_t2)
            latest["ВКО_Актуальный"] = latest["ВКО_Актуальный"].fillna(default_name)
            latest["Таб. номер ВКО_Актуальный"] = latest["Таб. номер ВКО_Актуальный"].fillna(default_id)
        else:
            # Для двух файлов: приоритет T-0 → T-1
            # Создаем DataFrame со всеми ключами из merged
            all_keys = merged[key_columns].drop_duplicates()
            
            # Переименовываем колонки для build_latest_manager
            # ВАЖНО: НЕ делаем merge с all_keys перед вызовом build_latest_manager,
            # потому что build_latest_manager сам делает outer join и объединяет все ключи
            best_current_renamed = best_current.rename(columns={"ВКО_T0": "ВКО", "Таб. номер ВКО_T0": "Таб. номер ВКО"})
            best_previous_renamed = best_previous.rename(columns={"ВКО_T1": "ВКО", "Таб. номер ВКО_T1": "Таб. номер ВКО"})
            
            # Вызываем build_latest_manager с исходными данными
            # build_latest_manager сам делает outer join и объединяет все ключи из обоих файлов
            latest = self.build_latest_manager(
                current_best=best_current_renamed,
                previous_best=best_previous_renamed,
                key_columns=key_columns,
                variant_name=variant_name,
            )
            
            # Убеждаемся, что latest содержит все ключи из merged
            # Если в latest нет ключа из merged, добавляем его с значениями по умолчанию
            latest = all_keys.merge(latest, on=key_columns, how="left")
            
            # Заполняем пропуски значениями по умолчанию ТОЛЬКО если менеджер не найден ни в одном файле
            # Если клиент есть только в T-0, то latest уже должен содержать правильное значение из T-0
            # после build_latest_manager, поэтому fillna применяется только к действительно отсутствующим значениям
            default_name = self.defaults["manager_name"]
            identifier_settings = self.identifiers["manager_id"]
            default_id = format_identifier(
                self.defaults["manager_id"],
                total_length=identifier_settings["total_length"],
                fill_char=identifier_settings["fill_char"],
            )
            # Заполняем только если действительно нет значения (не должно быть после build_latest_manager)
            latest["ВКО_Актуальный"] = latest["ВКО_Актуальный"].fillna(default_name)
            latest["Таб. номер ВКО_Актуальный"] = latest["Таб. номер ВКО_Актуальный"].fillna(default_id)
        
        merged = merged.merge(latest, on=key_columns, how="left")

        log_debug(
            self.logger,
            f"{variant_name}: итоговый набор содержит {len(merged)} строк",
            class_name="Aggregator",
            func_name="assemble_variant_dataset_with_t2",
        )
        return merged
    
    def build_manager_summary(
        self,
        variant_df: pd.DataFrame,
        include_tb: bool,
        summary_name: str,
        manager_columns: Mapping[str, str],
    ) -> pd.DataFrame:
        """Создаёт свод по уникальным ТН+ВКО (+ТБ опционально).
        
        Группирует данные по менеджеру (и ТБ, если include_tb=True) и суммирует
        факты и прирост.
        
        Args:
            variant_df: DataFrame с данными варианта
            include_tb: Если True, добавляет ТБ в группировку
            summary_name: Имя свода для логирования
            manager_columns: Словарь с именами колонок {"id": "Таб. номер ВКО_...", "name": "ВКО_..."}
        
        Returns:
            DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный), ТБ (если include_tb),
            Факт_T0, Факт_T1, Факт_T2 (если есть в variant_df), Прирост
        """
        manager_id_col = manager_columns["id"]
        manager_name_col = manager_columns["name"]

        group_columns = [manager_id_col, manager_name_col]
        tb_column_present = include_tb and "tb" in variant_df.columns
        if tb_column_present:
            group_columns.append("tb")

        # Определяем список числовых колонок для суммирования
        numeric_columns = ["Факт_T0", "Факт_T1", "Прирост"]
        # Добавляем Факт_T2, если он есть в данных (для варианта three)
        if "Факт_T2" in variant_df.columns:
            numeric_columns.insert(2, "Факт_T2")  # Вставляем между Факт_T1 и Прирост

        grouped = (
            variant_df.groupby(group_columns, dropna=False)[numeric_columns]
            .sum()
            .reset_index()
        )
        rename_map = {
            manager_id_col: SELECTED_MANAGER_ID_COL,
            manager_name_col: SELECTED_MANAGER_NAME_COL,
        }
        if tb_column_present:
            rename_map["tb"] = "ТБ"
        grouped = grouped.rename(columns=rename_map)

        log_debug(
            self.logger,
            f"{summary_name}: агрегировано {len(grouped)} строк",
            class_name="Aggregator",
            func_name="build_manager_summary",
        )
        return grouped


class VariantCalculator:
    """Базовый класс для расчета вариантов прироста.
    
    Определяет общий интерфейс для всех вариантов расчета.
    Каждый вариант должен реализовать метод calculate.
    
    Атрибуты:
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
        aggregator: Экземпляр Aggregator для агрегации данных
    """
    
    def __init__(
        self,
        defaults: Mapping[str, Any],
        identifiers: Mapping[str, Any],
        logger: Mapping[str, Any],
    ):
        """Инициализирует калькулятор варианта.
        
        Args:
            defaults: Словарь с настройками по умолчанию
            identifiers: Словарь с настройками форматирования идентификаторов
            logger: Логгер с методами info и debug
        """
        self.defaults = defaults
        self.identifiers = identifiers
        self.logger = logger
        self.aggregator = Aggregator(defaults, identifiers, logger)
    
    def calculate(
        self,
        current_df: pd.DataFrame,
        previous_df: pd.DataFrame,
        previous2_df: Optional[pd.DataFrame],
    ) -> pd.DataFrame:
        """Вычисляет вариант прироста.
        
        Должен быть переопределен в наследниках.
        
        Args:
            current_df: DataFrame с данными T-0
            previous_df: DataFrame с данными T-1
            previous2_df: DataFrame с данными T-2 (может быть None)
        
        Returns:
            DataFrame с результатами расчета варианта
        
        Raises:
            NotImplementedError: Если метод не переопределен
        """
        raise NotImplementedError("Метод calculate должен быть переопределен в наследнике")


class Variant1Calculator(VariantCalculator):
    """Калькулятор варианта 1: По КМ (manager_id), без учета ТБ.
    
    Логика расчета:
    - Суммируем в каждом файле все что на КМ по manager_id
    - Затем для каждого КМ вычитаем одно из другого
    - Если T-2 указан: прирост = (T-0 - T-1) - (T-1 - T-2)
    - Иначе: прирост = T-0 - T-1
    """
    
    def calculate(
        self,
        current_df: pd.DataFrame,
        previous_df: pd.DataFrame,
        previous2_df: Optional[pd.DataFrame],
    ) -> pd.DataFrame:
        """Вычисляет вариант 1: По КМ, без ТБ.
        
        Args:
            current_df: DataFrame с данными T-0
            previous_df: DataFrame с данными T-1
            previous2_df: DataFrame с данными T-2 (может быть None)
        
        Returns:
            DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный),
            Факт_T0, Факт_T1, Факт_T2 (если есть), Прирост
        """
        log_info(self.logger, "Расчет варианта 1: По КМ, без ТБ")
        
        # Агрегируем по manager_id в каждом файле
        agg_t0 = self.aggregator.aggregate_facts(current_df, ["manager_id"], "T0", "V1")
        agg_t1 = self.aggregator.aggregate_facts(previous_df, ["manager_id"], "T1", "V1")
        
        if previous2_df is not None:
            agg_t2 = self.aggregator.aggregate_facts(previous2_df, ["manager_id"], "T2", "V1")
            merged = (
                pd.merge(agg_t0, agg_t1, on=["manager_id"], how="outer")
                .merge(agg_t2, on=["manager_id"], how="outer")
                .fillna({"Факт_T0": 0.0, "Факт_T1": 0.0, "Факт_T2": 0.0})
            )
            merged["Прирост"] = (merged["Факт_T0"] - merged["Факт_T1"]) - (merged["Факт_T1"] - merged["Факт_T2"])
        else:
            merged = (
                pd.merge(agg_t0, agg_t1, on=["manager_id"], how="outer")
                .fillna({"Факт_T0": 0.0, "Факт_T1": 0.0})
            )
            merged["Прирост"] = merged["Факт_T0"] - merged["Факт_T1"]
        
        # Добавляем информацию о менеджере из исходных данных
        manager_info = pd.DataFrame(columns=["manager_id", "manager_name"])
        if not current_df.empty:
            manager_info = current_df[["manager_id", "manager_name"]].drop_duplicates()
        elif not previous_df.empty:
            manager_info = previous_df[["manager_id", "manager_name"]].drop_duplicates()
        elif previous2_df is not None and not previous2_df.empty:
            manager_info = previous2_df[["manager_id", "manager_name"]].drop_duplicates()
        
        if not manager_info.empty:
            result = merged.merge(manager_info, on=["manager_id"], how="left")
        else:
            result = merged.copy()
            result["manager_name"] = self.defaults.get("manager_name", "Не найден КМ")
        
        result = result.rename(columns={
            "manager_id": SELECTED_MANAGER_ID_COL,
            "manager_name": SELECTED_MANAGER_NAME_COL,
        })
        
        return result


class Variant2Calculator(VariantCalculator):
    """Калькулятор варианта 2: По ИНН (client_id), КМ определяется на конец без учета ТБ.
    
    Логика расчета:
    - Агрегация по client_id
    - КМ определяется приоритетом: T-0 → T-1 → T-2 (если T-2 указан)
    - Прирост рассчитывается с учетом T-2, если он указан
    """
    
    def calculate(
        self,
        current_df: pd.DataFrame,
        previous_df: pd.DataFrame,
        previous2_df: Optional[pd.DataFrame],
    ) -> pd.DataFrame:
        """Вычисляет вариант 2: По ИНН, без ТБ.
        
        Args:
            current_df: DataFrame с данными T-0
            previous_df: DataFrame с данными T-1
            previous2_df: DataFrame с данными T-2 (может быть None)
        
        Returns:
            DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный),
            Факт_T0, Факт_T1, Прирост
        """
        log_info(self.logger, "Расчет варианта 2: По ИНН, без ТБ")
        
        variant_df = self.aggregator.assemble_variant_dataset_with_t2(
            variant_name="V2_ИНН_безТБ",
            key_columns=["client_id"],
            current_df=current_df,
            previous_df=previous_df,
            previous2_df=previous2_df,
        )
        
        # Агрегируем по актуальному менеджеру
        summary = self.aggregator.build_manager_summary(
            variant_df=variant_df,
            include_tb=False,
            summary_name="V2_SUMMARY",
            manager_columns={"id": "Таб. номер ВКО_Актуальный", "name": "ВКО_Актуальный"},
        )
        
        return summary


class Variant3Calculator(VariantCalculator):
    """Калькулятор варианта 3: По ИНН (client_id), КМ определяется на конец с учетом ТБ.
    
    Логика расчета:
    - Агрегация по client_id и tb
    - КМ определяется приоритетом: T-0 → T-1 → T-2 (если T-2 указан)
    - При агрегации учитываем ТБ: собираем данные по клиенту только для КМ с тем же ТБ
    - Прирост рассчитывается с учетом T-2, если он указан
    """
    
    def calculate(
        self,
        current_df: pd.DataFrame,
        previous_df: pd.DataFrame,
        previous2_df: Optional[pd.DataFrame],
    ) -> pd.DataFrame:
        """Вычисляет вариант 3: По ИНН, с учетом ТБ.
        
        Args:
            current_df: DataFrame с данными T-0
            previous_df: DataFrame с данными T-1
            previous2_df: DataFrame с данными T-2 (может быть None)
        
        Returns:
            DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный), ТБ,
            Факт_T0, Факт_T1, Факт_T2 (если есть), Прирост
        """
        log_info(self.logger, "Расчет варианта 3: По ИНН, с учетом ТБ")
        
        variant_df = self.aggregator.assemble_variant_dataset_with_t2(
            variant_name="V3_ИНН_сТБ",
            key_columns=["client_id", "tb"],
            current_df=current_df,
            previous_df=previous_df,
            previous2_df=previous2_df,
        )
        
        # Агрегируем по актуальному менеджеру с учетом ТБ
        summary = self.aggregator.build_manager_summary(
            variant_df=variant_df,
            include_tb=True,
            summary_name="V3_SUMMARY",
            manager_columns={"id": "Таб. номер ВКО_Актуальный", "name": "ВКО_Актуальный"},
        )
        
        return summary


class PercentileCalculator:
    """Класс для расчета процентилей.
    
    Инкапсулирует логику расчета показателей "кого я обогнал" и "кто меня обогнал"
    как в общем разрезе, так и в разрезе ТерБанков и с условиями.
    
    Методы:
        append_percentile_columns: Добавляет колонки процентилей к таблице
    """
    
    @staticmethod
    def append_percentile_columns(
        table: pd.DataFrame,
        *,
        value_column: str,
        tb_column: Optional[str] = None,
        gosb_column: Optional[str] = None,
        percentile_filter: str = "all",
        group_by: str = "all",
    ) -> pd.DataFrame:
        """Добавляет в таблицу колонки процентных рангов и абсолютных значений.
        
        Добавляет следующие колонки:
        - Обогнал_всего_%: процент КМ с меньшим результатом (всего)
        - Обогнали_меня_всего_%: процент КМ с большим результатом (всего)
        - Обогнал_всего_кол: количество КМ с меньшим результатом
        - Обогнали_меня_всего_кол: количество КМ с большим результатом
        - Равных_всего_кол: количество КМ с таким же результатом
        - Всего_КМ_всего: общее количество КМ для расчета
        
        Расчет выполняется по данным выбранного варианта (с ТБ или без ТБ).
        Данные для расчета могут быть отфильтрованы по percentile_filter.
        Группировка может выполняться по ТБ, ГОСБ или их комбинации.
        
        Args:
            table: DataFrame с данными для расчета процентилей
            value_column: Имя колонки со значениями для расчета (например, "Прирост")
            tb_column: Имя колонки с ТБ для группировки (например, "ТБ"). 
                       Если None, расчеты по ТБ не выполняются
            gosb_column: Имя колонки с ГОСБ для группировки (например, "ГОСБ").
                        Если None, расчеты по ГОСБ не выполняются
            percentile_filter: Фильтр для данных при расчете процентилей ("all", ">=0", ">0" и т.д.)
            group_by: Тип группировки для расчета процентилей:
                     - "all" - среди всех КМ
                     - "tb" - среди КМ с тем же ТБ
                     - "gosb" - среди КМ с тем же ГОСБ
                     - "tb_and_gosb" - среди КМ с тем же ТБ и ГОСБ
        
        Returns:
            DataFrame с добавленными колонками процентилей
        
        Raises:
            KeyError: Если value_column не найдена в таблице
        """
        if value_column not in table.columns:
            raise KeyError(
                f"Колонка '{value_column}' не найдена в таблице для расчёта процентилей."
            )

        prepared = table.copy()
        values = pd.to_numeric(prepared[value_column], errors="coerce").fillna(0.0)

        # Применяем фильтр для расчета процентилей
        if percentile_filter and percentile_filter.lower() not in ("all", "все"):
            filter_mask = build_filter_mask(values, percentile_filter)
            filtered_indices = values[filter_mask].index
        else:
            filter_mask = pd.Series(True, index=values.index)
            filtered_indices = values.index
        
        # Рассчитываем процентили для каждой строки относительно отфильтрованного набора
        prepared["Обогнал_всего_%"] = 0.0
        prepared["Обогнали_меня_всего_%"] = 0.0
        prepared["Обогнал_всего_кол"] = 0
        prepared["Обогнали_меня_всего_кол"] = 0
        prepared["Равных_всего_кол"] = 0
        prepared["Всего_КМ_всего"] = 0
        
        # Для каждой строки рассчитываем процентили относительно отфильтрованного набора
        for idx in prepared.index:
            current_value = values.loc[idx]
            
            # Если значение не проходит фильтр, процентили = 0
            if percentile_filter and percentile_filter.lower() not in ("all", "все"):
                if not filter_mask.loc[idx]:
                    continue
            
            # Определяем набор для сравнения в зависимости от group_by
            comparison_indices = filtered_indices.copy()
            
            if group_by == "tb" and tb_column and tb_column in prepared.columns:
                # Сравниваем только с КМ того же ТБ
                current_tb = prepared.loc[idx, tb_column]
                tb_mask = prepared.loc[comparison_indices, tb_column] == current_tb
                comparison_indices = comparison_indices[tb_mask]
            elif group_by == "gosb" and gosb_column and gosb_column in prepared.columns:
                # Сравниваем только с КМ того же ГОСБ
                current_gosb = prepared.loc[idx, gosb_column]
                gosb_mask = prepared.loc[comparison_indices, gosb_column] == current_gosb
                comparison_indices = comparison_indices[gosb_mask]
            elif group_by == "tb_and_gosb":
                # Сравниваем только с КМ того же ТБ и ГОСБ
                if tb_column and tb_column in prepared.columns:
                    current_tb = prepared.loc[idx, tb_column]
                    tb_mask = prepared.loc[comparison_indices, tb_column] == current_tb
                    comparison_indices = comparison_indices[tb_mask]
                if gosb_column and gosb_column in prepared.columns:
                    current_gosb = prepared.loc[idx, gosb_column]
                    gosb_mask = prepared.loc[comparison_indices, gosb_column] == current_gosb
                    comparison_indices = comparison_indices[gosb_mask]
            # Если group_by == "all", используем все filtered_indices
            
            # Исключаем текущую строку из сравнения
            comparison_indices = comparison_indices[comparison_indices != idx]
            
            if len(comparison_indices) == 0:
                continue
            
            comparison_values = values.loc[comparison_indices]
            
            # Считаем процентили для текущего значения относительно группы сравнения
            less_count = (comparison_values < current_value).sum()
            greater_count = (comparison_values > current_value).sum()
            equal_count = (comparison_values == current_value).sum()
            
            total = len(comparison_indices)
            if total > 0:
                prepared.loc[idx, "Обогнал_всего_%"] = round((less_count / total) * 100, 2)
                prepared.loc[idx, "Обогнали_меня_всего_%"] = round((greater_count / total) * 100, 2)
            prepared.loc[idx, "Обогнал_всего_кол"] = less_count
            prepared.loc[idx, "Обогнали_меня_всего_кол"] = greater_count
            prepared.loc[idx, "Равных_всего_кол"] = max(0, equal_count)
            prepared.loc[idx, "Всего_КМ_всего"] = total

        return prepared


class ExcelExporter:
    """Класс для экспорта данных в Excel с форматированием.
    
    Инкапсулирует логику записи DataFrame в Excel файлы с применением
    форматирования: ширина колонок, выравнивание, числовые форматы.
    
    Методы:
        format_sheet: Применяет форматирование к листу Excel
        write_sheet: Записывает DataFrame в лист Excel с форматированием
    """
    
    @staticmethod
    def format_sheet(
        writer: pd.ExcelWriter, 
        sheet_name: str, 
        df: pd.DataFrame,
        min_width: int = 20,
        max_width: int = 200,
        wrap_text: bool = True,
    ) -> None:
        """Применяет форматирование листа Excel через openpyxl.
        
        Выполняет следующие операции:
        1. Замораживает первую строку (заголовки)
        2. Включает автофильтр
        3. Форматирует заголовки (жирный шрифт, перенос текста)
        4. Настраивает ширину колонок автоматически по содержимому (с ограничениями min_width-max_width)
        5. Включает перенос текста для всех ячеек (если wrap_text=True)
        6. Применяет числовые форматы:
           - #,##0.00 для процентов и фактов
           - #,##0 для количеств
        
        Args:
            writer: ExcelWriter для записи
            sheet_name: Имя листа для форматирования
            df: DataFrame с данными (используется для определения типов колонок)
            min_width: Минимальная ширина колонки в пунктах (по умолчанию 20)
            max_width: Максимальная ширина колонки в пунктах (по умолчанию 200)
            wrap_text: Включить перенос текста по строкам (по умолчанию True)
        """
        workbook = writer.book
        worksheet = workbook[sheet_name]

        if df.empty:
            return

        worksheet.freeze_panes = worksheet["A2"]
        worksheet.auto_filter.ref = worksheet.dimensions

        # Настройки выравнивания с учетом wrap_text
        header_alignment = Alignment(wrap_text=wrap_text, vertical="top")
        wrap_alignment = Alignment(wrap_text=wrap_text, vertical="top")
        number_alignment = Alignment(wrap_text=wrap_text, vertical="top")
        header_font = Font(bold=True)

        # Форматируем заголовки
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
            cell.font = header_font
            cell.alignment = header_alignment

        # Автоматическая подстройка ширины колонок по содержимому
        for col_idx, column in enumerate(df.columns, start=1):
            # Собираем все значения в колонке (заголовок + данные)
            values = [str(column)] + [str(value) for value in df[column].tolist()]
            
            # Находим максимальную длину содержимого
            max_len = max((len(str(value)) for value in values), default=0)
            
            # Добавляем небольшой отступ (2 символа) для комфортного отображения
            calculated_width = max_len + 2
            
            # Применяем ограничения min_width и max_width
            width = clamp_width(calculated_width, min_width, max_width)
            
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width

            # Форматируем данные в колонке
            if worksheet.max_row >= 2:
                data_range = worksheet[f"{column_letter}2": f"{column_letter}{worksheet.max_row}"]
                if (
                    column.startswith("Факт")
                    or column == "Прирост"
                    or "Обогнал" in column
                    or "Обогнали" in column
                    or column == "FACT_VALUE"
                    or column == "PLAN_VALUE"
                    or column == "Факт"
                ):
                    for cell_tuple in data_range:
                        for item in cell_tuple:
                            item.number_format = "#,##0.00"
                            item.alignment = number_alignment
                elif "_кол" in column or "Всего_КМ" in column or "Кол-во" in column:
                    # Колонки с количеством - целые числа
                    for cell_tuple in data_range:
                        for item in cell_tuple:
                            item.number_format = "#,##0"
                            item.alignment = number_alignment
                else:
                    for cell_tuple in data_range:
                        for item in cell_tuple:
                            item.alignment = wrap_alignment
    
    @staticmethod
    def write_sheet(
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        written_sheets: Set[str],
        min_width: int = 20,
        max_width: int = 200,
        wrap_text: bool = True,
    ) -> None:
        """Записывает DataFrame в лист Excel с форматированием.
        
        Проверяет, не был ли лист уже записан, и если нет - записывает данные
        и применяет форматирование.
        
        Args:
            writer: ExcelWriter для записи
            sheet_name: Имя листа для записи
            df: DataFrame с данными
            written_sheets: Множество уже записанных листов (изменяется на месте)
            min_width: Минимальная ширина колонки в пунктах (по умолчанию 20)
            max_width: Максимальная ширина колонки в пунктах (по умолчанию 200)
            wrap_text: Включить перенос текста по строкам (по умолчанию True)
        """
        if sheet_name in written_sheets:
            return
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ExcelExporter.format_sheet(writer, sheet_name, df, min_width, max_width, wrap_text)
        written_sheets.add(sheet_name)


def build_logger(log_dir: Path, topic: str) -> Dict[str, Any]:
    """Инициализирует файловый логгер и возвращает набор функций."""

    ensure_directories([log_dir])
    suffix = timestamp_suffix()
    info_path = log_dir / f"INFO_{topic}{suffix}.log"
    debug_path = log_dir / f"DEBUG_{topic}{suffix}.log"

    # INFO всегда дублируется в консоль, DEBUG пишется только в файл (согласно ТЗ).
    def info(message: str) -> None:
        line = f"{dt.datetime.now():%Y-%m-%d %H:%M:%S} - [INFO] - {message}"
        print(line)
        with info_path.open("a", encoding="utf-8") as info_file:
            info_file.write(f"{line}\n")

    def debug(message: str, class_name: str, func_name: str) -> None:
        line = (
            f"{dt.datetime.now():%Y-%m-%d %H:%M:%S} - [DEBUG] - "
            f"{message} [class: {class_name} | def: {func_name}]"
        )
        with debug_path.open("a", encoding="utf-8") as debug_file:
            debug_file.write(f"{line}\n")

    return {"info": info, "debug": debug}


def log_info(logger: Mapping[str, Any], message: str) -> None:
    """Упрощает вызов INFO-сообщений."""

    logger["info"](message)


def log_debug(
    logger: Mapping[str, Any], message: str, class_name: str, func_name: str
) -> None:
    """Упрощает вызов DEBUG-сообщений."""

    logger["debug"](message, class_name, func_name)


# -------------------------- Работа с исходными файлами ----------------------


def read_source_file(
    file_path: Path,
    sheet_name: str,
    columns: List[Dict[str, str]],
    drop_rules: Mapping[str, Iterable[str]],
    identifiers: Mapping[str, Mapping[str, Any]],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Загружает исходный Excel и подготавливает данные.
    
    Функция-обертка для DataLoader.read_source_file.
    Сохранена для обратной совместимости.
    
    Args:
        file_path: Путь к файлу Excel
        sheet_name: Имя листа для чтения
        columns: Список словарей с alias и source для колонок
        drop_rules: Правила фильтрации строк
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с очищенными и отформатированными данными
    """
    data_loader = DataLoader(identifiers, logger)
    return data_loader.read_source_file(file_path, sheet_name, columns, drop_rules)


def drop_forbidden_rows(
    df: pd.DataFrame,
    drop_rules: Mapping[str, Iterable[str]],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Удаляет строки с запрещёнными значениями.
    
    Функция-обертка для DataLoader.drop_forbidden_rows.
    Сохранена для обратной совместимости.
    
    Args:
        df: DataFrame для очистки
        drop_rules: Словарь {column_alias: tuple(forbidden_values)}
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame без запрещенных строк
    """
    # Создаем временный загрузчик (identifiers не нужны для drop_forbidden_rows)
    identifiers = {"manager_id": {"total_length": 8, "fill_char": "0"}, "client_id": {"total_length": 12, "fill_char": "0"}}
    data_loader = DataLoader(identifiers, logger)
    return data_loader.drop_forbidden_rows(df, drop_rules)


# -------------------------- Агрегация данных --------------------------------


def aggregate_facts(
    df: pd.DataFrame,
    key_columns: List[str],
    suffix: str,
    logger: Mapping[str, Any],
    variant_name: str,
) -> pd.DataFrame:
    """Группирует данные по ключу и суммирует факт.
    
    Функция-обертка для Aggregator.aggregate_facts.
    Сохранена для обратной совместимости.
    
    Args:
        df: Исходный DataFrame с данными
        key_columns: Список колонок для группировки
        suffix: Суффикс для имени результирующей колонки
        logger: Логгер для записи сообщений
        variant_name: Имя варианта для логирования
    
    Returns:
        DataFrame с колонками key_columns и Факт_{suffix}
    """
    # Создаем временный агрегатор (defaults и identifiers не нужны для aggregate_facts)
    defaults = {"manager_name": "", "manager_id": ""}
    identifiers = {"manager_id": {"total_length": 8, "fill_char": "0"}, "client_id": {"total_length": 12, "fill_char": "0"}}
    aggregator = Aggregator(defaults, identifiers, logger)
    return aggregator.aggregate_facts(df, key_columns, suffix, variant_name)


def select_best_manager(
    df: pd.DataFrame,
    key_columns: List[str],
    logger: Mapping[str, Any],
    variant_name: str,
) -> pd.DataFrame:
    """Определяет доминантного менеджера (по сумме факта) для каждого ключа.
    
    Функция-обертка для Aggregator.select_best_manager.
    Сохранена для обратной совместимости.
    
    Args:
        df: Исходный DataFrame с данными
        key_columns: Список колонок для ключа
        logger: Логгер для записи сообщений
        variant_name: Имя варианта для логирования
    
    Returns:
        DataFrame с колонками key_columns, "ВКО", "Таб. номер ВКО"
    """
    defaults = {"manager_name": "", "manager_id": ""}
    identifiers = {"manager_id": {"total_length": 8, "fill_char": "0"}, "client_id": {"total_length": 12, "fill_char": "0"}}
    aggregator = Aggregator(defaults, identifiers, logger)
    return aggregator.select_best_manager(df, key_columns, variant_name)


def build_latest_manager(
    current_best: pd.DataFrame,
    previous_best: pd.DataFrame,
    key_columns: List[str],
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
    variant_name: str,
) -> pd.DataFrame:
    """Комбинирует менеджеров, отдавая приоритет файлу T-0.
    
    Функция-обертка для Aggregator.build_latest_manager.
    Сохранена для обратной совместимости.
    
    Args:
        current_best: DataFrame с менеджерами из T-0
        previous_best: DataFrame с менеджерами из T-1
        key_columns: Список колонок для ключа
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
        variant_name: Имя варианта для логирования
    
    Returns:
        DataFrame с колонками key_columns, "ВКО_Актуальный", "Таб. номер ВКО_Актуальный"
    """
    aggregator = Aggregator(defaults, identifiers, logger)
    return aggregator.build_latest_manager(current_best, previous_best, key_columns, variant_name)


def assemble_variant_dataset(
    variant_name: str,
    key_columns: List[str],
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Формирует таблицу для конкретного варианта ключа."""

    log_debug(
        logger,
        f"{variant_name}: старт построения набора данных",
        class_name="Aggregator",
        func_name="assemble_variant_dataset",
    )

    agg_current = aggregate_facts(current_df, key_columns, "T0", logger, variant_name)
    agg_previous = aggregate_facts(previous_df, key_columns, "T1", logger, variant_name)
    merged = (
        pd.merge(agg_current, agg_previous, on=key_columns, how="outer")
        .fillna({"Факт_T0": 0.0, "Факт_T1": 0.0})
    )
    merged["Прирост"] = merged["Факт_T0"] - merged["Факт_T1"]

    best_current = select_best_manager(
        current_df, key_columns, logger, variant_name
    ).rename(columns={"ВКО": "ВКО_T0", "Таб. номер ВКО": "Таб. номер ВКО_T0"})
    best_previous = select_best_manager(
        previous_df, key_columns, logger, variant_name
    ).rename(columns={"ВКО": "ВКО_T1", "Таб. номер ВКО": "Таб. номер ВКО_T1"})

    merged = merged.merge(best_current, on=key_columns, how="left")
    merged = merged.merge(best_previous, on=key_columns, how="left")

    latest = build_latest_manager(
        current_best=best_current.rename(columns={"ВКО_T0": "ВКО", "Таб. номер ВКО_T0": "Таб. номер ВКО"}),
        previous_best=best_previous.rename(columns={"ВКО_T1": "ВКО", "Таб. номер ВКО_T1": "Таб. номер ВКО"}),
        key_columns=key_columns,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
        variant_name=variant_name,
    )
    merged = merged.merge(latest, on=key_columns, how="left")

    log_debug(
        logger,
        f"{variant_name}: итоговый набор содержит {len(merged)} строк",
        class_name="Aggregator",
        func_name="assemble_variant_dataset",
    )
    return merged


def build_manager_summary(
    variant_df: pd.DataFrame,
    include_tb: bool,
    logger: Mapping[str, Any],
    summary_name: str,
    manager_columns: Mapping[str, str],
) -> pd.DataFrame:
    """Создаёт свод по уникальным ТН+ВКО (+ТБ опционально).
    
    Функция-обертка для Aggregator.build_manager_summary.
    Сохранена для обратной совместимости.
    
    Args:
        variant_df: DataFrame с данными варианта
        include_tb: Если True, добавляет ТБ в группировку
        logger: Логгер для записи сообщений
        summary_name: Имя свода для логирования
        manager_columns: Словарь с именами колонок {"id": "...", "name": "..."}
    
    Returns:
        DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный), ТБ (если include_tb),
        Факт_T0, Факт_T1, Прирост
    """
    defaults = {"manager_name": "", "manager_id": ""}
    identifiers = {"manager_id": {"total_length": 8, "fill_char": "0"}, "client_id": {"total_length": 12, "fill_char": "0"}}
    aggregator = Aggregator(defaults, identifiers, logger)
    return aggregator.build_manager_summary(variant_df, include_tb, summary_name, manager_columns)


def clamp_width(length: int, min_width: int = 20, max_width: int = 200) -> int:
    """Ограничивает ширину столбца в заданном диапазоне.
    
    Args:
        length: Требуемая ширина столбца
        min_width: Минимальная ширина (по умолчанию 20)
        max_width: Максимальная ширина (по умолчанию 200)
    
    Returns:
        Ширина столбца в пределах [min_width, max_width]
    """
    return max(min_width, min(length, max_width))


def format_excel_sheet(
    writer: pd.ExcelWriter, 
    sheet_name: str, 
    df: pd.DataFrame,
    min_width: int = 20,
    max_width: int = 200,
    wrap_text: bool = True,
) -> None:
    """Применяет форматирование листа Excel через openpyxl.
    
    Функция-обертка для ExcelExporter.format_sheet.
    Сохранена для обратной совместимости.
    
    Args:
        writer: ExcelWriter для записи
        sheet_name: Имя листа для форматирования
        df: DataFrame с данными
        min_width: Минимальная ширина колонки в пунктах (по умолчанию 20)
        max_width: Максимальная ширина колонки в пунктах (по умолчанию 200)
        wrap_text: Включить перенос текста по строкам (по умолчанию True)
    """
    ExcelExporter.format_sheet(writer, sheet_name, df, min_width, max_width, wrap_text)


def format_decimal_string(value: float, decimals: int = 5) -> str:
    """Форматирует число вида 0.00000."""

    numeric_value = 0.0 if value is None or pd.isna(value) else float(value)
    return f"{numeric_value:.{decimals}f}"


def build_spod_dataset(
    source_table: pd.DataFrame,
    *,
    value_column: str,
    fact_value_filter: str,
    plan_value: float,
    priority: int,
    contest_code: str,
    tournament_code: str,
    contest_date: str,
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
    dataset_name: str,
    percentile_value_column: Optional[str] = None,
) -> pd.DataFrame:
    """Готовит данные для загрузки в СПОД.
    
    Args:
        source_table: Исходная таблица с данными
        value_column: Колонка для фильтрации и сортировки
        fact_value_filter: Фильтр для отбора данных
        plan_value: Плановое значение
        priority: Приоритет
        contest_code: Код конкурса
        tournament_code: Код турнира
        contest_date: Дата конкурса
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер
        dataset_name: Имя датасета для логирования
        percentile_value_column: Колонка для FACT_VALUE (если отличается от value_column).
                                 Если None, используется value_column.
    """

    if value_column not in source_table.columns:
        raise KeyError(
            f"Колонка '{value_column}' отсутствует в источнике '{dataset_name}'."
        )

    # Определяем колонку для FACT_VALUE
    fact_value_column = percentile_value_column if percentile_value_column is not None else value_column
    
    if fact_value_column not in source_table.columns:
        raise KeyError(
            f"Колонка '{fact_value_column}' для FACT_VALUE отсутствует в источнике '{dataset_name}'."
        )

    mask = build_filter_mask(source_table[value_column], fact_value_filter)
    filtered = source_table[mask].copy()
    filtered = filtered.sort_values(by=value_column, ascending=False)

    log_debug(
        logger,
        f"SPOD '{dataset_name}': после фильтра {fact_value_filter} осталось {len(filtered)} строк",
        class_name="Exporter",
        func_name="build_spod_dataset",
    )

    dataset = filtered.rename(
        columns={
            SELECTED_MANAGER_ID_COL: "MANAGER_PERSON_NUMBER",
        }
    )["MANAGER_PERSON_NUMBER"].to_frame()

    manager_identifier = identifiers["manager_id"]
    dataset["MANAGER_PERSON_NUMBER"] = dataset["MANAGER_PERSON_NUMBER"].apply(
        lambda value: format_identifier(
            value=value,
            total_length=max(manager_identifier["total_length"], 20),
            fill_char=manager_identifier["fill_char"],
        )
    )
    dataset["CONTEST_CODE"] = contest_code
    dataset["TOURNAMENT_CODE"] = tournament_code
    dataset["CONTEST_DATE"] = parse_contest_date(contest_date)
    dataset["PLAN_VALUE"] = format_decimal_string(plan_value)
    # Используем fact_value_column для FACT_VALUE
    dataset["FACT_VALUE"] = filtered[fact_value_column].apply(format_decimal_string)
    dataset["priority_type"] = priority

    log_debug(
        logger,
        f"SPOD '{dataset_name}': подготовлено {len(dataset)} строк для выгрузки",
        class_name="Exporter",
        func_name="build_spod_dataset",
    )

    return dataset[
        [
            "MANAGER_PERSON_NUMBER",
            "CONTEST_CODE",
            "TOURNAMENT_CODE",
            "CONTEST_DATE",
            "PLAN_VALUE",
            "FACT_VALUE",
            "priority_type",
        ]
    ]


def build_spod_dataset_for_excel(
    source_table: pd.DataFrame,
    filtered_table: pd.DataFrame,
    spod_dataset: pd.DataFrame,
    value_column: str,
    source_type: str,
    manager_tb_mapping: pd.Series,
    manager_gosb_mapping: pd.Series,
    variant_df_for_client_summary: Optional[pd.DataFrame],
    current_df: Optional[pd.DataFrame],
    previous_df: Optional[pd.DataFrame],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Создает расширенный SPOD датасет для Excel с дополнительными колонками.
    
    Добавляет к базовому SPOD датасету:
    - Факт (форматированный: целая часть с разделителями, дробная 2 знака)
    - ФИО КМ
    - ТБ
    - ГОСБ
    - Количество ИНН (для вариантов по ИНН)
    - Для процентильного SPOD: кол-во КМ кого обогнал, кто обогнал, равных, всего
    
    Args:
        source_table: Исходная таблица (summary_tn или percentile_tn)
        filtered_table: Отфильтрованная таблица (после применения fact_value_filter)
        spod_dataset: Базовый SPOD датасет
        value_column: Колонка со значениями
        manager_tb_mapping: Маппинг табельного номера на ТБ
        manager_gosb_mapping: Маппинг табельного номера на ГОСБ
        variant_df_for_client_summary: variant_df для подсчета ИНН (для вариантов 2 и 3)
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с дополнительными колонками для Excel
    """
    result = spod_dataset.copy()
    
    # Создаем маппинги из filtered_table по табельному номеру
    # Форматируем табельные номера в filtered_table для сопоставления с MANAGER_PERSON_NUMBER
    manager_identifier = identifiers.get("manager_id", {"total_length": 8, "fill_char": "0"})
    
    # Форматируем табельные номера в filtered_table так же, как в build_spod_dataset
    filtered_table_mapped = filtered_table.copy()
    filtered_table_mapped["MANAGER_PERSON_NUMBER_FORMATTED"] = filtered_table_mapped[SELECTED_MANAGER_ID_COL].apply(
        lambda value: format_identifier(
            value=value,
            total_length=max(manager_identifier.get("total_length", 8), 20),
            fill_char=manager_identifier.get("fill_char", "0"),
        )
    )
    
    # Создаем маппинги по отформатированному табельному номеру
    manager_name_map = filtered_table_mapped[[
        "MANAGER_PERSON_NUMBER_FORMATTED", SELECTED_MANAGER_NAME_COL
    ]].drop_duplicates().set_index("MANAGER_PERSON_NUMBER_FORMATTED")[SELECTED_MANAGER_NAME_COL]
    
    fact_values_map = filtered_table_mapped[[
        "MANAGER_PERSON_NUMBER_FORMATTED", value_column
    ]].drop_duplicates().set_index("MANAGER_PERSON_NUMBER_FORMATTED")[value_column]
    
    # Добавляем ФИО КМ
    result["ФИО КМ"] = result["MANAGER_PERSON_NUMBER"].map(manager_name_map).fillna("")
    
    # Добавляем ТБ и ГОСБ (используем исходные табельные номера из filtered_table)
    # Создаем маппинг отформатированных номеров к исходным
    formatted_to_original = filtered_table_mapped.set_index("MANAGER_PERSON_NUMBER_FORMATTED")[
        SELECTED_MANAGER_ID_COL
    ].to_dict()
    
    # Получаем исходные табельные номера для маппинга ТБ и ГОСБ
    result["_original_manager_id"] = result["MANAGER_PERSON_NUMBER"].map(formatted_to_original).fillna(
        result["MANAGER_PERSON_NUMBER"]
    )
    result["ТБ"] = result["_original_manager_id"].map(manager_tb_mapping).fillna("")
    result["ГОСБ"] = result["_original_manager_id"].map(manager_gosb_mapping).fillna("")
    result = result.drop(columns=["_original_manager_id"])
    
    # Добавляем Факт (число в числовом формате, будет отформатировано в Excel как #,##0.00)
    result["Факт"] = result["MANAGER_PERSON_NUMBER"].map(fact_values_map).fillna(0.0)
    
    # Добавляем количество ИНН
    if variant_df_for_client_summary is not None:
        # Для вариантов 2 и 3 (по ИНН) - подсчитываем количество уникальных ИНН для каждого менеджера
        if "Таб. номер ВКО_Актуальный" in variant_df_for_client_summary.columns:
            inn_count = variant_df_for_client_summary.groupby("Таб. номер ВКО_Актуальный")["client_id"].nunique()
            # Форматируем табельные номера для сопоставления
            manager_identifier = identifiers.get("manager_id", {"total_length": 8, "fill_char": "0"})
            inn_count_formatted = {}
            for orig_id, count in inn_count.items():
                formatted_id = format_identifier(orig_id, max(manager_identifier.get("total_length", 8), 20), 
                                                 manager_identifier.get("fill_char", "0"))
                inn_count_formatted[formatted_id] = count
            result["Кол-во ИНН"] = result["MANAGER_PERSON_NUMBER"].map(inn_count_formatted).fillna(0).astype(int)
        else:
            result["Кол-во ИНН"] = 0
    elif current_df is not None:
        # Для варианта 1 (по КМ) - считаем количество уникальных ИНН из исходных данных (T-0 и T-1)
        manager_identifier = identifiers.get("manager_id", {"total_length": 8, "fill_char": "0"})
        
        # Объединяем T-0 и T-1 для подсчета всех ИНН
        if previous_df is not None:
            combined_df = pd.concat([current_df[["manager_id", "client_id"]], 
                                    previous_df[["manager_id", "client_id"]]])
        else:
            combined_df = current_df[["manager_id", "client_id"]]
        
        inn_count = combined_df.groupby("manager_id")["client_id"].nunique()
        inn_count_formatted = {}
        for orig_id, count in inn_count.items():
            formatted_id = format_identifier(orig_id, max(manager_identifier.get("total_length", 8), 20), 
                                             manager_identifier.get("fill_char", "0"))
            inn_count_formatted[formatted_id] = count
        result["Кол-во ИНН"] = result["MANAGER_PERSON_NUMBER"].map(inn_count_formatted).fillna(0).astype(int)
    else:
        result["Кол-во ИНН"] = 0
    
    # Для процентильного SPOD добавляем колонки с количеством (только для scenario_percentile)
    if source_type == "scenario_percentile":
        # Используем source_table (percentile_tn) для получения процентилей, так как filtered_table может не содержать все строки
        # Но маппим по отфильтрованным табельным номерам из filtered_table_mapped
        if "Обогнал_всего_кол" in source_table.columns:
            # Форматируем табельные номера в source_table для сопоставления
            source_table_mapped = source_table.copy()
            source_table_mapped["MANAGER_PERSON_NUMBER_FORMATTED"] = source_table_mapped[SELECTED_MANAGER_ID_COL].apply(
                lambda value: format_identifier(
                    value=value,
                    total_length=max(manager_identifier.get("total_length", 8), 20),
                    fill_char=manager_identifier.get("fill_char", "0"),
                )
            )
            
            # Создаем маппинги по отформатированному табельному номеру из source_table
            percentile_count_map = source_table_mapped.set_index("MANAGER_PERSON_NUMBER_FORMATTED")[
                ["Обогнал_всего_кол", "Обогнали_меня_всего_кол", "Равных_всего_кол", "Всего_КМ_всего"]
            ]
            
            result["Обогнал_всего_кол"] = result["MANAGER_PERSON_NUMBER"].map(
                percentile_count_map["Обогнал_всего_кол"]
            ).fillna(0).astype(int)
            result["Обогнали_меня_всего_кол"] = result["MANAGER_PERSON_NUMBER"].map(
                percentile_count_map["Обогнали_меня_всего_кол"]
            ).fillna(0).astype(int)
            result["Равных_всего_кол"] = result["MANAGER_PERSON_NUMBER"].map(
                percentile_count_map["Равных_всего_кол"]
            ).fillna(0).astype(int)
            result["Всего_КМ_всего"] = result["MANAGER_PERSON_NUMBER"].map(
                percentile_count_map["Всего_КМ_всего"]
            ).fillna(0).astype(int)
    
    # Переупорядочиваем колонки: сначала стандартные SPOD, потом дополнительные
    base_cols = [
        "MANAGER_PERSON_NUMBER",
        "CONTEST_CODE",
        "TOURNAMENT_CODE",
        "CONTEST_DATE",
        "PLAN_VALUE",
        "FACT_VALUE",
        "priority_type",
    ]
    additional_cols = ["Факт", "ФИО КМ", "ТБ", "ГОСБ", "Кол-во ИНН"]
    percentile_cols = ["Обогнал_всего_кол", "Обогнали_меня_всего_кол", "Равных_всего_кол", "Всего_КМ_всего"]
    
    # Оставляем только существующие колонки
    existing_base = [col for col in base_cols if col in result.columns]
    existing_additional = [col for col in additional_cols if col in result.columns]
    
    # Процентильные колонки добавляем только для scenario_percentile
    if source_type == "scenario_percentile":
        existing_percentile = [col for col in percentile_cols if col in result.columns]
        result = result[existing_base + existing_additional + existing_percentile]
    else:
        result = result[existing_base + existing_additional]
    
    return result


def rename_output_columns(
    df: pd.DataFrame, alias_to_source: Mapping[str, str]
) -> pd.DataFrame:
    """Возвращает DataFrame с русскими заголовками для ключей."""

    renamed = df.copy()
    # Переименовываем лишь ключевые идентификаторы; остальные остаются в машинном виде.
    mapping = {
        alias: alias_to_source.get(alias, alias)
        for alias in ("client_id", "tb", "manager_id")
    }
    renamed = renamed.rename(columns=mapping)
    return renamed


def format_raw_sheet(
    df: pd.DataFrame,
    alias_to_source: Mapping[str, str],
) -> pd.DataFrame:
    """Возвращает DataFrame для исходного листа с читаемыми колонками и типами."""

    printable = df.copy()

    # Переименовываем только те столбцы, которые известны пользователю.
    rename_mapping = {
        alias: alias_to_source.get(alias, alias)
        for alias in df.columns
        if alias in alias_to_source
    }
    printable = printable.rename(columns=rename_mapping)

    # Числовой факт выводим отдельным столбцом с гарантированным float.
    if "fact_value_clean" in printable.columns:
        printable = printable.rename(
            columns={"fact_value_clean": "Факт (число)"}
        )
        printable["Факт (число)"] = printable["Факт (число)"].apply(
            lambda value: 0.0 if value is None or pd.isna(value) else float(value)
        )

    return printable


def build_direct_manager_summary(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    include_tb: bool,
    logger: Mapping[str, Any],
    summary_name: str,
) -> pd.DataFrame:
    """Суммирует факты по менеджерам напрямую (без ключа клиента)."""

    base_columns = ["manager_id", "manager_name"]
    tb_column_present = include_tb and "tb" in current_df.columns and "tb" in previous_df.columns
    if tb_column_present:
        base_columns.append("tb")

    def aggregate(source_df: pd.DataFrame, suffix: str) -> pd.DataFrame:
        grouped = (
            source_df[base_columns + ["fact_value_clean"]]
            .fillna({"fact_value_clean": 0.0})
            .groupby(base_columns, dropna=False, as_index=False)
            .sum(numeric_only=True)
        )
        return grouped.rename(columns={"fact_value_clean": f"Факт_{suffix}"})

    agg_current = aggregate(current_df, "T0")
    agg_previous = aggregate(previous_df, "T1")

    merged = pd.merge(agg_current, agg_previous, on=base_columns, how="outer").fillna(
        {"Факт_T0": 0.0, "Факт_T1": 0.0}
    )
    merged["Прирост"] = merged["Факт_T0"] - merged["Факт_T1"]

    rename_map = {
        "manager_id": DIRECT_MANAGER_ID_COL,
        "manager_name": DIRECT_MANAGER_NAME_COL,
    }
    if tb_column_present:
        rename_map["tb"] = "ТБ"
    result = merged.rename(columns=rename_map)

    log_debug(
        logger,
        f"{summary_name}: агрегировано {len(result)} строк (прямой подсчёт по менеджеру)",
        class_name="Aggregator",
        func_name="build_direct_manager_summary",
    )
    return result


def build_latest_manager_with_t2(
    current_best: pd.DataFrame,
    previous_best: pd.DataFrame,
    previous2_best: Optional[pd.DataFrame],
    key_columns: List[str],
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
    variant_name: str,
) -> pd.DataFrame:
    """Комбинирует менеджеров, отдавая приоритет файлу T-0, затем T-1, затем T-2.
    
    Функция-обертка для Aggregator.build_latest_manager_with_t2.
    Сохранена для обратной совместимости.
    
    Args:
        current_best: DataFrame с менеджерами из T-0
        previous_best: DataFrame с менеджерами из T-1
        previous2_best: DataFrame с менеджерами из T-2 (может быть None)
        key_columns: Список колонок для ключа
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
        variant_name: Имя варианта для логирования
    
    Returns:
        DataFrame с колонками key_columns, "ВКО_Актуальный", "Таб. номер ВКО_Актуальный"
    """
    aggregator = Aggregator(defaults, identifiers, logger)
    return aggregator.build_latest_manager_with_t2(current_best, previous_best, previous2_best, key_columns, variant_name)


def assemble_variant_dataset_with_t2(
    variant_name: str,
    key_columns: List[str],
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    previous2_df: Optional[pd.DataFrame],
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Формирует таблицу для конкретного варианта ключа с поддержкой T-2.
    
    Функция-обертка для Aggregator.assemble_variant_dataset_with_t2.
    Сохранена для обратной совместимости.
    
    Args:
        variant_name: Имя варианта для логирования
        key_columns: Список колонок для ключа агрегации
        current_df: DataFrame с данными T-0
        previous_df: DataFrame с данными T-1
        previous2_df: DataFrame с данными T-2 (может быть None)
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с колонками key_columns, Факт_T0, Факт_T1, Факт_T2 (если есть),
        Прирост, ВКО_T0, ВКО_T1, ВКО_T2 (если есть), ВКО_Актуальный, Таб. номер ВКО_Актуальный
    """
    aggregator = Aggregator(defaults, identifiers, logger)
    return aggregator.assemble_variant_dataset_with_t2(
        variant_name, key_columns, current_df, previous_df, previous2_df
    )


# Функции-обертки для обратной совместимости (используют классы внутри)
def calculate_variant_1(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    previous2_df: Optional[pd.DataFrame],
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Вариант 1: По КМ (manager_id), без учета ТБ.
    
    Функция-обертка для Variant1Calculator.
    
    Args:
        current_df: DataFrame с данными T-0
        previous_df: DataFrame с данными T-1
        previous2_df: DataFrame с данными T-2 (может быть None)
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с результатами варианта 1
    """
    calculator = Variant1Calculator(defaults, identifiers, logger)
    return calculator.calculate(current_df, previous_df, previous2_df)


def calculate_variant_2(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    previous2_df: Optional[pd.DataFrame],
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Вариант 2: По ИНН (client_id), КМ определяется на конец без учета ТБ.
    
    Функция-обертка для Variant2Calculator.
    
    Args:
        current_df: DataFrame с данными T-0
        previous_df: DataFrame с данными T-1
        previous2_df: DataFrame с данными T-2 (может быть None)
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с результатами варианта 2
    """
    calculator = Variant2Calculator(defaults, identifiers, logger)
    return calculator.calculate(current_df, previous_df, previous2_df)


def calculate_single_file_count(
    df: pd.DataFrame,
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Расчет по одному файлу: количество строк (сделок) для каждого ТН.
    
    Для каждого табельного номера считается количество строк в файле.
    
    Args:
        df: DataFrame с данными одного файла
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный), Количество_сделок
    """
    log_info(logger, "Расчет по одному файлу: количество сделок для каждого ТН")
    
    # Группируем по manager_id и считаем количество строк
    result = df.groupby("manager_id", as_index=False).agg({
        "manager_name": "first",  # Берем первое значение ФИО
        "fact_value_clean": "count",  # Считаем количество строк
    }).rename(columns={
        "manager_id": SELECTED_MANAGER_ID_COL,
        "manager_name": SELECTED_MANAGER_NAME_COL,
        "fact_value_clean": "Количество_сделок",
    })
    
    log_info(logger, f"Найдено {len(result)} уникальных ТН, всего сделок: {result['Количество_сделок'].sum()}")
    return result


def calculate_single_file_max(
    df: pd.DataFrame,
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Расчет по одному файлу: максимальная сумма среди строк для каждого КМ.
    
    Для каждого табельного номера находится максимальная сумма fact_value_clean среди всех строк.
    
    Args:
        df: DataFrame с данными одного файла
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный), Максимальная_сумма
    """
    log_info(logger, "Расчет по одному файлу: максимальная сумма для каждого КМ")
    
    # Группируем по manager_id и находим максимальную сумму
    result = df.groupby("manager_id", as_index=False).agg({
        "manager_name": "first",  # Берем первое значение ФИО
        "fact_value_clean": "max",  # Находим максимальную сумму
    }).rename(columns={
        "manager_id": SELECTED_MANAGER_ID_COL,
        "manager_name": SELECTED_MANAGER_NAME_COL,
        "fact_value_clean": "Максимальная_сумма",
    })
    
    log_info(logger, f"Найдено {len(result)} уникальных ТН")
    return result


def calculate_new_clients(
    files_2025: List[pd.DataFrame],
    files_2024: List[pd.DataFrame],
    key_mode: str,
    include_tb: bool,
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Расчет новых клиентов: ИНН с суммой факта в 2025 > 0, но сумма факта в 2024 = 0 или его нет.
    
    Логика:
    - Для каждого ИНН суммируем факт по всем файлам 2025 года
    - Для каждого ИНН суммируем факт по всем файлам 2024 года
    - Находим ИНН, где сумма 2025 > 0 и сумма 2024 = 0 или его нет
    - Агрегируем по ТН (с учетом ТБ, если нужно)
    - Выводим: сумма 2024, сумма 2025, количество месяцев с суммой в 2024, количество месяцев с суммой в 2025
    
    Args:
        files_2025: Список DataFrame с данными файлов 2025 года (12 файлов)
        files_2024: Список DataFrame с данными файлов 2024 года (12 файлов)
        key_mode: Режим агрегации ("manager" или "client")
        include_tb: Учитывать ли ТБ при расчете
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с колонками: Таб. номер ВКО (выбранный), ВКО (выбранный), ТБ (если include_tb),
        Сумма_2024, Сумма_2025, Месяцев_с_суммой_2024, Месяцев_с_суммой_2025
    """
    log_info(logger, "Расчет новых клиентов: поиск ИНН с фактом в 2025, но без факта в 2024")
    
    # Определяем ключи для агрегации
    if key_mode == "client":
        if include_tb:
            agg_keys = ["client_id", "tb"]
        else:
            agg_keys = ["client_id"]
    else:
        agg_keys = ["manager_id"]
    
    # Для определения итогового ТН нужно объединить файлы в правильном порядке:
    # сначала 2025 от декабря к январю (обратный порядок), потом 2024 от декабря к январю (обратный порядок)
    # Это нужно для того, чтобы при использовании "last" в агрегации брался последний менеджер из этой последовательности
    files_for_manager_selection = []
    
    # Добавляем файлы 2025 года в обратном порядке (от декабря к январю)
    if files_2025:
        files_for_manager_selection.extend(reversed(files_2025))
    
    # Добавляем файлы 2024 года в обратном порядке (от декабря к январю)
    if files_2024:
        files_for_manager_selection.extend(reversed(files_2024))
    
    # Объединяем все файлы для определения итогового ТН
    if files_for_manager_selection:
        df_all_for_manager = pd.concat(files_for_manager_selection, ignore_index=True)
    else:
        df_all_for_manager = pd.DataFrame()
    
    # Объединяем все файлы 2025 года (в обычном порядке для подсчета сумм)
    if files_2025:
        df_2025_all = pd.concat(files_2025, ignore_index=True)
    else:
        df_2025_all = pd.DataFrame()
    
    # Объединяем все файлы 2024 года (в обычном порядке для подсчета сумм)
    if files_2024:
        df_2024_all = pd.concat(files_2024, ignore_index=True)
    else:
        df_2024_all = pd.DataFrame()
    
    # Агрегируем по ИНН (или ТН) для 2025 года
    if not df_2025_all.empty:
        # Считаем сумму факта для каждого ИНН
        agg_dict = {
            "fact_value_clean": "sum",
        }
        if key_mode == "client":
            # Для определения итогового ТН используем объединенный DataFrame всех файлов
            # Выбираем ТН с максимальной суммой факта по каждому ИНН
            if not df_all_for_manager.empty:
                # Группируем по ИНН и ТН, суммируем факт
                grouping_cols = agg_keys + ["manager_id", "manager_name"]
                if include_tb and "tb" not in agg_keys:
                    grouping_cols.append("tb")
                
                grouped = (
                    df_all_for_manager[grouping_cols + ["fact_value_clean"]]
                    .fillna({"fact_value_clean": 0.0})
                    .groupby(grouping_cols, dropna=False, as_index=False)
                    .sum(numeric_only=True)
                )
                
                # Для каждого ИНН выбираем ТН с максимальной суммой факта
                # Если суммы равны, idxmax вернет первый из равных
                idx = grouped.groupby(agg_keys, dropna=False)["fact_value_clean"].idxmax()
                manager_agg = grouped.loc[idx, grouping_cols].copy()
                
                # Убираем fact_value_clean из результата, оставляем только нужные колонки
                result_cols = agg_keys + ["manager_id", "manager_name"]
                if include_tb:
                    if "tb" in manager_agg.columns:
                        result_cols.append("tb")
                manager_agg = manager_agg[result_cols].copy()
            else:
                manager_agg = pd.DataFrame(columns=agg_keys + ["manager_id", "manager_name"])
                if include_tb:
                    manager_agg["tb"] = []
        else:
            # Для key_mode == "manager" добавляем manager_name в агрегацию
            agg_dict["manager_name"] = "last"
        
        agg_2025 = df_2025_all.groupby(agg_keys, as_index=False).agg(agg_dict)
        agg_2025["Сумма_2025"] = agg_2025["fact_value_clean"]
        
        # Добавляем итоговый ТН для каждого ИНН (только для key_mode == "client")
        if key_mode == "client" and not df_all_for_manager.empty:
            agg_2025 = pd.merge(agg_2025, manager_agg, on=agg_keys, how="left")
        
        # Считаем количество месяцев с суммой > 0 для каждого ИНН
        # Для этого нужно посчитать по каждому файлу отдельно
        months_with_sum_2025 = []
        for df_file in files_2025:
            if not df_file.empty:
                file_agg = df_file.groupby(agg_keys, as_index=False).agg({"fact_value_clean": "sum"})
                file_agg["has_sum"] = (file_agg["fact_value_clean"] > 0).astype(int)
                months_with_sum_2025.append(file_agg[agg_keys + ["has_sum"]])
        
        if months_with_sum_2025:
            months_df = pd.concat(months_with_sum_2025, ignore_index=True)
            months_count = months_df.groupby(agg_keys, as_index=False).agg({"has_sum": "sum"})
            months_count = months_count.rename(columns={"has_sum": "Месяцев_с_суммой_2025"})
            agg_2025 = pd.merge(agg_2025, months_count, on=agg_keys, how="left")
            agg_2025["Месяцев_с_суммой_2025"] = agg_2025["Месяцев_с_суммой_2025"].fillna(0).astype(int)
        else:
            agg_2025["Месяцев_с_суммой_2025"] = 0
        
        agg_2025 = agg_2025.drop(columns=["fact_value_clean"])
    else:
        agg_2025 = pd.DataFrame(columns=agg_keys + ["Сумма_2025", "Месяцев_с_суммой_2025"])
        if key_mode == "client":
            agg_2025["manager_id"] = []
            agg_2025["manager_name"] = []
            if include_tb:
                agg_2025["tb"] = []
    
    # Агрегируем по ИНН (или ТН) для 2024 года
    if not df_2024_all.empty:
        # Считаем сумму факта для каждого ИНН
        agg_dict_2024 = {
            "fact_value_clean": "sum",
        }
        if key_mode == "manager":
            # Для key_mode == "manager" добавляем manager_name в агрегацию
            agg_dict_2024["manager_name"] = "last"
        agg_2024 = df_2024_all.groupby(agg_keys, as_index=False).agg(agg_dict_2024)
        agg_2024["Сумма_2024"] = agg_2024["fact_value_clean"]
        
        # Считаем количество месяцев с суммой > 0 для каждого ИНН
        months_with_sum_2024 = []
        for df_file in files_2024:
            if not df_file.empty:
                file_agg = df_file.groupby(agg_keys, as_index=False).agg({"fact_value_clean": "sum"})
                file_agg["has_sum"] = (file_agg["fact_value_clean"] > 0).astype(int)
                months_with_sum_2024.append(file_agg[agg_keys + ["has_sum"]])
        
        if months_with_sum_2024:
            months_df = pd.concat(months_with_sum_2024, ignore_index=True)
            months_count = months_df.groupby(agg_keys, as_index=False).agg({"has_sum": "sum"})
            months_count = months_count.rename(columns={"has_sum": "Месяцев_с_суммой_2024"})
            agg_2024 = pd.merge(agg_2024, months_count, on=agg_keys, how="left")
            agg_2024["Месяцев_с_суммой_2024"] = agg_2024["Месяцев_с_суммой_2024"].fillna(0).astype(int)
        else:
            agg_2024["Месяцев_с_суммой_2024"] = 0
        
        agg_2024 = agg_2024.drop(columns=["fact_value_clean"])
    else:
        agg_2024 = pd.DataFrame(columns=agg_keys + ["Сумма_2024", "Месяцев_с_суммой_2024"])
    
    # Объединяем данные 2025 и 2024
    if key_mode == "client":
        # Для client_id объединяем по ИНН
        merged = pd.merge(agg_2025, agg_2024, on=agg_keys, how="left", suffixes=("", "_2024"))
        merged["Сумма_2024"] = merged["Сумма_2024"].fillna(0.0)
        merged["Месяцев_с_суммой_2024"] = merged["Месяцев_с_суммой_2024"].fillna(0)
        
        # Фильтруем: сумма 2025 > 0 и сумма 2024 = 0
        new_clients = merged[(merged["Сумма_2025"] > 0) & (merged["Сумма_2024"] == 0)].copy()
        
        # Агрегируем по ТН (с учетом ТБ, если нужно)
        if include_tb:
            group_keys = ["manager_id", "tb"]
        else:
            group_keys = ["manager_id"]
        
        result = new_clients.groupby(group_keys, as_index=False).agg({
            "manager_name": "first",
            "Сумма_2024": "sum",
            "Сумма_2025": "sum",
            "Месяцев_с_суммой_2024": "sum",
            "Месяцев_с_суммой_2025": "sum",
        })
        
        rename_map = {
            "manager_id": SELECTED_MANAGER_ID_COL,
            "manager_name": SELECTED_MANAGER_NAME_COL,
        }
        if include_tb:
            rename_map["tb"] = "ТБ"
        result = result.rename(columns=rename_map)
    else:
        # Для manager_id объединяем по ТН
        merged = pd.merge(agg_2025, agg_2024, on=agg_keys, how="left", suffixes=("", "_2024"))
        merged["Сумма_2024"] = merged["Сумма_2024"].fillna(0.0)
        merged["Месяцев_с_суммой_2024"] = merged["Месяцев_с_суммой_2024"].fillna(0)
        
        # Фильтруем: сумма 2025 > 0 и сумма 2024 = 0
        result = merged[(merged["Сумма_2025"] > 0) & (merged["Сумма_2024"] == 0)].copy()
        
        rename_map = {
            "manager_id": SELECTED_MANAGER_ID_COL,
            "manager_name": SELECTED_MANAGER_NAME_COL,
        }
        result = result.rename(columns=rename_map)
    
    log_info(logger, f"Найдено {len(result)} уникальных ТН с новыми клиентами")
    return result


def calculate_variant_3(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    previous2_df: Optional[pd.DataFrame],
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> pd.DataFrame:
    """Вариант 3: По ИНН (client_id), КМ определяется на конец с учетом ТБ.
    
    Функция-обертка для Variant3Calculator.
    
    Args:
        current_df: DataFrame с данными T-0
        previous_df: DataFrame с данными T-1
        previous2_df: DataFrame с данными T-2 (может быть None)
        defaults: Настройки по умолчанию
        identifiers: Настройки форматирования идентификаторов
        logger: Логгер для записи сообщений
    
    Returns:
        DataFrame с результатами варианта 3
    """
    calculator = Variant3Calculator(defaults, identifiers, logger)
    return calculator.calculate(current_df, previous_df, previous2_df)


def build_variant_matrix(
    current_df: pd.DataFrame,
    previous_df: pd.DataFrame,
    defaults: Mapping[str, Any],
    identifiers: Mapping[str, Any],
    logger: Mapping[str, Any],
) -> Dict[int, pd.DataFrame]:
    """Строит матрицу всех 8 вариантов расчета приростов согласно методологии.
    
    Возвращает словарь {номер_варианта: DataFrame} для всех 8 комбинаций:
    1: ВКО, без ТБ, КМ по каждому файлу
    2: ВКО, с ТБ, КМ по каждому файлу
    3: ВКО, без ТБ, последний КМ
    4: ВКО, с ТБ, последний КМ
    5: ИНН, без ТБ, КМ по каждому файлу
    6: ИНН, с ТБ, КМ по каждому файлу
    7: ИНН, без ТБ, последний КМ
    8: ИНН, с ТБ, последний КМ
    """
    
    results: Dict[int, pd.DataFrame] = {}
    
    # Варианты 5-8: ИНН (client_id) - используем существующие функции
    log_info(logger, "Строю варианты 5-8: ИНН")
    
    # Вариант 5: ИНН, без ТБ, КМ по каждому файлу
    log_info(logger, "Строю вариант 5: ИНН, без ТБ, КМ по каждому файлу")
    variant_5_df = assemble_variant_dataset(
        variant_name="V5_ИНН_безТБ_КМ_пофайлу",
        key_columns=["client_id"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    # Агрегируем по КМ из T0 (доминантный в каждом файле)
    summary_5 = build_manager_summary(
        variant_df=variant_5_df,
        include_tb=False,
        logger=logger,
        summary_name="V5_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_T0", "name": "ВКО_T0"},
    )
    results[5] = summary_5
    
    # Вариант 6: ИНН, с ТБ, КМ по каждому файлу
    log_info(logger, "Строю вариант 6: ИНН, с ТБ, КМ по каждому файлу")
    variant_6_df = assemble_variant_dataset(
        variant_name="V6_ИНН_сТБ_КМ_пофайлу",
        key_columns=["client_id", "tb"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    summary_6 = build_manager_summary(
        variant_df=variant_6_df,
        include_tb=True,
        logger=logger,
        summary_name="V6_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_T0", "name": "ВКО_T0"},
    )
    results[6] = summary_6
    
    # Вариант 7: ИНН, без ТБ, последний КМ
    log_info(logger, "Строю вариант 7: ИНН, без ТБ, последний КМ")
    variant_7_df = assemble_variant_dataset(
        variant_name="V7_ИНН_безТБ_КМ_последний",
        key_columns=["client_id"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    summary_7 = build_manager_summary(
        variant_df=variant_7_df,
        include_tb=False,
        logger=logger,
        summary_name="V7_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_Актуальный", "name": "ВКО_Актуальный"},
    )
    results[7] = summary_7
    
    # Вариант 8: ИНН, с ТБ, последний КМ
    log_info(logger, "Строю вариант 8: ИНН, с ТБ, последний КМ")
    variant_8_df = assemble_variant_dataset(
        variant_name="V8_ИНН_сТБ_КМ_последний",
        key_columns=["client_id", "tb"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    summary_8 = build_manager_summary(
        variant_df=variant_8_df,
        include_tb=True,
        logger=logger,
        summary_name="V8_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_Актуальный", "name": "ВКО_Актуальный"},
    )
    results[8] = summary_8
    
    # Варианты 1-4: ВКО (gosb) - аналогичная логика, но с gosb как ключом
    log_info(logger, "Строю варианты 1-4: ВКО")
    
    # Вариант 1: ВКО, без ТБ, КМ по каждому файлу
    log_info(logger, "Строю вариант 1: ВКО, без ТБ, КМ по каждому файлу")
    variant_1_df = assemble_variant_dataset(
        variant_name="V1_ВКО_безТБ_КМ_пофайлу",
        key_columns=["gosb"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    summary_1 = build_manager_summary(
        variant_df=variant_1_df,
        include_tb=False,
        logger=logger,
        summary_name="V1_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_T0", "name": "ВКО_T0"},
    )
    results[1] = summary_1
    
    # Вариант 2: ВКО, с ТБ, КМ по каждому файлу
    log_info(logger, "Строю вариант 2: ВКО, с ТБ, КМ по каждому файлу")
    variant_2_df = assemble_variant_dataset(
        variant_name="V2_ВКО_сТБ_КМ_пофайлу",
        key_columns=["gosb", "tb"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    summary_2 = build_manager_summary(
        variant_df=variant_2_df,
        include_tb=True,
        logger=logger,
        summary_name="V2_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_T0", "name": "ВКО_T0"},
    )
    results[2] = summary_2
    
    # Вариант 3: ВКО, без ТБ, последний КМ
    log_info(logger, "Строю вариант 3: ВКО, без ТБ, последний КМ")
    variant_3_df = assemble_variant_dataset(
        variant_name="V3_ВКО_безТБ_КМ_последний",
        key_columns=["gosb"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    summary_3 = build_manager_summary(
        variant_df=variant_3_df,
        include_tb=False,
        logger=logger,
        summary_name="V3_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_Актуальный", "name": "ВКО_Актуальный"},
    )
    results[3] = summary_3
    
    # Вариант 4: ВКО, с ТБ, последний КМ
    log_info(logger, "Строю вариант 4: ВКО, с ТБ, последний КМ")
    variant_4_df = assemble_variant_dataset(
        variant_name="V4_ВКО_сТБ_КМ_последний",
        key_columns=["gosb", "tb"],
        current_df=current_df,
        previous_df=previous_df,
        defaults=defaults,
        identifiers=identifiers,
        logger=logger,
    )
    summary_4 = build_manager_summary(
        variant_df=variant_4_df,
        include_tb=True,
        logger=logger,
        summary_name="V4_SUMMARY",
        manager_columns={"id": "Таб. номер ВКО_Актуальный", "name": "ВКО_Актуальный"},
    )
    results[4] = summary_4
    
    return results


# ----------------------------- Основной сценарий ----------------------------


def process_project(project_root: Path) -> None:
    """Запускает полный цикл обработки данных.
    
    Основная функция-оркестратор, которая:
    1. Загружает настройки проекта
    2. Инициализирует компоненты (DataLoader, VariantCalculator, PercentileCalculator, ExcelExporter)
    3. Загружает исходные файлы (T-0, T-1, T-2 если указан)
    4. Рассчитывает 3 варианта прироста
    5. Добавляет процентили для каждого варианта
    6. Экспортирует результаты в Excel
    
    Args:
        project_root: Корневая директория проекта (содержит IN/, OUT/, log/)
    """
    # 1. Собираем настройки (файлы, фильтры, идентификаторы, листы Excel)
    settings = build_settings_tree()
    file_section = settings["files"]
    defaults = settings["defaults"]
    identifiers = settings["identifiers"]
    spod_config = settings["spod"]
    report_layout = settings.get("report_layout", {})
    excel_formatting = settings.get("excel_formatting", {})
    
    # Получаем параметры форматирования Excel
    column_width_config = excel_formatting.get("column_width", {})
    min_width = column_width_config.get("min_width", 20)
    max_width = column_width_config.get("max_width", 200)
    wrap_text = excel_formatting.get("wrap_text", True)

    def build_whitelist(key: str) -> Optional[Set[str]]:
        """Возвращает множество разрешённых листов для указанного блока."""

        values = report_layout.get(key)
        if values is None:
            return None
        return set(values)

    detail_sheet_whitelist = build_whitelist("detail_sheets")
    summary_sheet_whitelist = build_whitelist("summary_sheets")
    spod_variant_whitelist = build_whitelist("spod_variants")
    raw_sheet_whitelist = build_whitelist("raw_sheets")

    # Готовим быстрый индекс по ключам файлов (current / previous / previous2).
    file_index = {item["key"]: item for item in file_section["items"]}
    current_meta = file_index["current"]
    previous_meta = file_index["previous"]
    previous2_meta = file_index.get("previous2")
    
    # Получаем параметры основного расчета для определения количества файлов
    main_calc_config = settings.get("main_calculation", {})
    use_files_count = main_calc_config.get("use_files_count", "two")
    
    if use_files_count not in ["one", "two", "three", "new"]:
        error_msg = f"Некорректное значение use_files_count: {use_files_count}. Допустимые значения: 'one', 'two', 'three' или 'new'"
        print(f"ОШИБКА: {error_msg}")
        log_info(logger, error_msg)
        return
    
    # Определяем необходимость использования T-2 на основе параметра
    use_t2 = (use_files_count == "three")

    input_dir = project_root / "IN"
    output_dir = project_root / "OUT"
    log_dir = project_root / "log"
    ensure_directories([input_dir, output_dir, log_dir])

    logger = build_logger(log_dir, spod_config["log_topic"])
    log_info(logger, "Старт обработки проекта YEAR_SPOD_Active_Rost_Ost")

    def should_write(entity_name: str, whitelist: Optional[Set[str]], block_name: str) -> bool:
        """Проверяет необходимость выгрузки листа согласно report_layout."""

        if whitelist is None:
            return True
        if entity_name in whitelist:
            return True
        log_debug(
            logger,
            f"Элемент '{entity_name}' пропущен (report_layout ограничил блок '{block_name}')",
            class_name="Exporter",
            func_name="process_project",
        )
        return False

    try:
        # Получаем колонки и фильтры для каждого файла (только для режимов "one", "two", "three")
        if use_files_count != "new":
            current_columns = get_file_columns(file_section, "current", defaults)
            current_filters = get_file_filters(file_section, "current", defaults)
            current_drop_rules = build_drop_rules(current_filters.get("drop_rules", []))
            current_column_profiles = build_column_profiles(current_columns)
            current_rename_map = current_column_profiles["rename_map"]
            current_alias_to_source = current_column_profiles["alias_to_source"]
            
            previous_columns = get_file_columns(file_section, "previous", defaults)
            previous_filters = get_file_filters(file_section, "previous", defaults)
            previous_drop_rules = build_drop_rules(previous_filters.get("drop_rules", []))
            previous_column_profiles = build_column_profiles(previous_columns)
            previous_rename_map = previous_column_profiles["rename_map"]
            previous_alias_to_source = previous_column_profiles["alias_to_source"]
            
            current_file = input_dir / current_meta["file_name"]
            previous_file = input_dir / previous_meta["file_name"]
            sheet_current = resolve_sheet_name(file_section, "current")
            sheet_previous = resolve_sheet_name(file_section, "previous")
        else:
            # Для режима "new" эти переменные не используются
            current_columns = []
            current_filters = {}
            current_drop_rules = []
            current_column_profiles = {}
            current_rename_map = {}
            current_alias_to_source = {}
            previous_columns = []
            previous_filters = {}
            previous_drop_rules = []
            previous_column_profiles = {}
            previous_rename_map = {}
            previous_alias_to_source = {}
            current_file = None
            previous_file = None
            sheet_current = None
            sheet_previous = None

        # Проверяем наличие необходимых файлов в зависимости от use_files_count
        missing_files = []
        
        if use_files_count == "one":
            # Для одного файла проверяем single_file настройки
            single_file_config = main_calc_config.get("single_file", {})
            single_file_name = single_file_config.get("file_name", "").strip()
            if not single_file_name:
                error_msg = "Для use_files_count='one' необходимо указать single_file.file_name в настройках"
                print(f"ОШИБКА: {error_msg}")
                log_info(logger, error_msg)
                return
            
            single_file_path = input_dir / single_file_name
            if not single_file_path.exists():
                missing_files.append(f"Один файл (single_file): {single_file_name}")
        elif use_files_count in ["two", "three"]:
            # Для двух и трех файлов проверяем current и previous
            if not current_file.exists():
                missing_files.append(f"T-0 (current): {current_meta['file_name']}")
            if not previous_file.exists():
                missing_files.append(f"T-1 (previous): {previous_meta['file_name']}")
            
            if use_files_count == "three":
                if previous2_meta is None:
                    missing_files.append("T-2 (previous2): не указан в настройках")
                else:
                    previous2_file = input_dir / previous2_meta["file_name"]
                    if not previous2_file.exists():
                        missing_files.append(f"T-2 (previous2): {previous2_meta['file_name']}")
        # Для режима "new" проверка файлов не выполняется здесь, файлы проверяются при загрузке
        
        if missing_files:
            error_msg = (
                f"Невозможно выполнить расчет: отсутствуют необходимые файлы для расчета по {use_files_count} файлам.\n"
                f"Отсутствующие файлы:\n" + "\n".join(f"  - {f}" for f in missing_files)
            )
            print(f"ОШИБКА: {error_msg}")
            log_info(logger, error_msg)
            return

        # Инициализируем загрузчик данных
        data_loader = DataLoader(identifiers, logger)
        
        # Получаем параметры основного расчета и процентиля (use_files_count уже получен выше)
        percentile_calc_config = settings.get("percentile_calculation", {})
        
        # Проверяем наличие необходимых блоков настроек
        if not main_calc_config:
            raise ValueError("Блок 'main_calculation' не найден в настройках")
        if not percentile_calc_config:
            raise ValueError("Блок 'percentile_calculation' не найден в настройках")
        
        # Обработка одного файла
        if use_files_count == "one":
            # Получаем параметры для одного файла
            one_file_config = main_calc_config.get("one_file", {})
            calculation_type = one_file_config.get("calculation_type", "count")
            
            # Получаем настройки из files.items с key="single" (НЕ используем defaults)
            single_meta = get_file_meta(file_section, "single")
            single_file_name = single_meta.get("file_name", "").strip()
            
            if not single_file_name:
                error_msg = "Для use_files_count='one' необходимо указать file_name в files.items с key='single'"
                print(f"ОШИБКА: {error_msg}")
                log_info(logger, error_msg)
                return
            
            single_file_columns = get_file_columns(file_section, "single", defaults, use_defaults=False)
            single_file_filters = get_file_filters(file_section, "single", defaults, use_defaults=False)
            single_file_sheet = resolve_sheet_name(file_section, "single")
            single_file_drop_rules = build_drop_rules(single_file_filters.get("drop_rules", []))
            single_file_in_rules = single_file_filters.get("in_rules", [])
            
            single_file_path = input_dir / single_file_name
            
            # Загружаем один файл
            single_df = data_loader.read_source_file(
                single_file_path,
                single_file_sheet,
                single_file_columns,
                single_file_drop_rules,
            )
            
            # Применяем IN фильтры
            if single_file_in_rules:
                single_df = data_loader.apply_in_rules(single_df, single_file_in_rules)
                log_info(logger, f"После применения IN фильтров осталось строк: {len(single_df)}")
            
            if calculation_type == "count":
                # Количество строк (сделок) для каждого ТН
                selected_summary = calculate_single_file_count(
                    single_df, defaults, identifiers, logger
                )
                value_column = "Количество_сделок"
            elif calculation_type == "max":
                # Максимальная сумма среди строк для каждого КМ
                selected_summary = calculate_single_file_max(
                    single_df, defaults, identifiers, logger
                )
                value_column = "Максимальная_сумма"
            else:
                raise ValueError(f"Неизвестный calculation_type: {calculation_type}. Допустимые значения: 'count' или 'max'")
            
            tb_column = None
            variant_df_for_client_summary = None
            current_df = single_df  # Для маппинга ТБ и ГОСБ
            previous_df = pd.DataFrame()  # Пустой для маппинга
            
        elif use_files_count == "new":
            # Загружаем 24 файла (12 для 2025 и 12 для 2024)
            files_2025 = []
            files_2024 = []
            
            # Получаем параметры для нового варианта
            new_files_config = main_calc_config.get("new_files", {})
            key_mode = new_files_config.get("key_mode", "client")
            include_tb = new_files_config.get("include_tb", False)
            
            log_info(logger, f"Загрузка файлов для нового варианта расчета: key_mode={key_mode}, include_tb={include_tb}")
            
            # Загружаем файлы 2025 года (от января к декабрю: M-01, M-02, ..., M-12)
            for i in range(12):
                month_num = i + 1
                file_key = f"2025_M-{month_num:02d}"
                try:
                    file_meta = get_file_meta(file_section, file_key)
                    file_name = file_meta.get("file_name", "").strip()
                    if file_name:
                        file_path = input_dir / file_name
                        if file_path.exists():
                            file_columns = get_file_columns(file_section, file_key, defaults)
                            file_filters = get_file_filters(file_section, file_key, defaults)
                            file_drop_rules = build_drop_rules(file_filters.get("drop_rules", []))
                            file_sheet = resolve_sheet_name(file_section, file_key)
                            
                            df = data_loader.read_source_file(
                                file_path,
                                file_sheet,
                                file_columns,
                                file_drop_rules,
                            )
                            files_2025.append(df)
                            log_info(logger, f"Загружен файл 2025_M-{month_num:02d}: {file_name}")
                        else:
                            log_info(logger, f"Файл 2025_M-{month_num:02d} не найден: {file_name}, пропускаем")
                    else:
                        log_info(logger, f"Имя файла для 2025_M-{month_num:02d} не указано, пропускаем")
                except KeyError:
                    log_info(logger, f"Конфигурация для 2025_M-{month_num:02d} не найдена, пропускаем")
            
            # Загружаем файлы 2024 года (от января к декабрю: M-01, M-02, ..., M-12)
            for i in range(12):
                month_num = i + 1
                file_key = f"2024_M-{month_num:02d}"
                try:
                    file_meta = get_file_meta(file_section, file_key)
                    file_name = file_meta.get("file_name", "").strip()
                    if file_name:
                        file_path = input_dir / file_name
                        if file_path.exists():
                            file_columns = get_file_columns(file_section, file_key, defaults)
                            file_filters = get_file_filters(file_section, file_key, defaults)
                            file_drop_rules = build_drop_rules(file_filters.get("drop_rules", []))
                            file_sheet = resolve_sheet_name(file_section, file_key)
                            
                            df = data_loader.read_source_file(
                                file_path,
                                file_sheet,
                                file_columns,
                                file_drop_rules,
                            )
                            files_2024.append(df)
                            log_info(logger, f"Загружен файл 2024_M-{month_num:02d}: {file_name}")
                        else:
                            log_info(logger, f"Файл 2024_M-{month_num:02d} не найден: {file_name}, пропускаем")
                    else:
                        log_info(logger, f"Имя файла для 2024_M-{month_num:02d} не указано, пропускаем")
                except KeyError:
                    log_info(logger, f"Конфигурация для 2024_M-{month_num:02d} не найдена, пропускаем")
            
            log_info(logger, f"Загружено файлов 2025: {len(files_2025)}, файлов 2024: {len(files_2024)}")
            
            # Рассчитываем новых клиентов
            selected_summary = calculate_new_clients(
                files_2025,
                files_2024,
                key_mode,
                include_tb,
                defaults,
                identifiers,
                logger,
            )
            
            value_column = "Сумма_2025"
            tb_column = "ТБ" if include_tb else None
            variant_df_for_client_summary = None
            
            # Для маппинга ТБ и ГОСБ используем первый доступный файл 2025
            if files_2025:
                current_df = files_2025[0]
            else:
                current_df = pd.DataFrame()
            previous_df = pd.DataFrame()  # Пустой для маппинга
            
        else:
            # Загружаем файлы T-0 и T-1
            current_df = data_loader.read_source_file(
                current_file,
                sheet_current,
                current_columns,
                current_drop_rules,
            )
            previous_df = data_loader.read_source_file(
                previous_file,
                sheet_previous,
                previous_columns,
                previous_drop_rules,
            )
            
            # Загружаем T-2, если требуется
            previous2_df = None
            if use_t2:
                previous2_file = input_dir / previous2_meta["file_name"]
                previous2_columns = get_file_columns(file_section, "previous2", defaults)
                previous2_filters = get_file_filters(file_section, "previous2", defaults)
                previous2_drop_rules = build_drop_rules(previous2_filters.get("drop_rules", []))
                previous2_df = data_loader.read_source_file(
                    previous2_file,
                    resolve_sheet_name(file_section, "previous2"),
                    previous2_columns,
                    previous2_drop_rules,
                )
                log_info(logger, f"Загружен файл T-2: {previous2_meta['file_name']}")
            
            # Получаем параметры основного расчета в зависимости от количества файлов
            if use_files_count == "two":
                two_files_config = main_calc_config.get("two_files", {})
                key_mode = two_files_config.get("key_mode", "client")
                include_tb = two_files_config.get("include_tb", False)
            elif use_files_count == "three":
                three_files_config = main_calc_config.get("three_files", {})
                key_mode = three_files_config.get("key_mode", "client")
                include_tb = three_files_config.get("include_tb", False)
            elif use_files_count == "new":
                new_files_config = main_calc_config.get("new_files", {})
                key_mode = new_files_config.get("key_mode", "client")
                include_tb = new_files_config.get("include_tb", False)
            else:
                raise ValueError(f"Неизвестный use_files_count: {use_files_count}")
            
            log_info(logger, f"Параметры основного расчета: key_mode={key_mode}, include_tb={include_tb}")
            
            # Рассчитываем основной свод в зависимости от параметров
            variant_df_for_client_summary = None
            
            if key_mode == "manager":
                # Расчет по КМ (manager_id), без учета ТБ
                selected_summary = calculate_variant_1(
                    current_df, previous_df, previous2_df if use_t2 else None,
                    defaults, identifiers, logger
                )
                tb_column = None
            elif key_mode == "client":
                if include_tb:
                    # Расчет по ИНН (client_id), с учетом ТБ
                    aggregator = Aggregator(defaults, identifiers, logger)
                    variant_df_for_client_summary = aggregator.assemble_variant_dataset_with_t2(
                        variant_name="ИНН_сТБ",
                        key_columns=["client_id", "tb"],
                        current_df=current_df,
                        previous_df=previous_df,
                        previous2_df=previous2_df if use_t2 else None,
                    )
                    selected_summary = calculate_variant_3(
                        current_df, previous_df, previous2_df if use_t2 else None,
                        defaults, identifiers, logger
                    )
                    tb_column = "ТБ"
                else:
                    # Расчет по ИНН (client_id), без учета ТБ
                    aggregator = Aggregator(defaults, identifiers, logger)
                    variant_df_for_client_summary = aggregator.assemble_variant_dataset_with_t2(
                        variant_name="ИНН_безТБ",
                        key_columns=["client_id"],
                        current_df=current_df,
                        previous_df=previous_df,
                        previous2_df=previous2_df if use_t2 else None,
                    )
                    selected_summary = calculate_variant_2(
                        current_df, previous_df, previous2_df if use_t2 else None,
                        defaults, identifiers, logger
                    )
                    tb_column = None
            else:
                raise ValueError(f"Неизвестный key_mode: {key_mode}. Допустимые значения: 'manager' или 'client'")
            
            value_column = "Прирост"
        
        # Объединяем SUMMARY_TN и PERCENTILE_TN в один лист
        # Сначала данные по расчету приростов, затем процентили
        summary_tn_combined = selected_summary.copy()
        
        # Добавляем ТБ и ГОСБ для каждого табельного номера (нужно для расчета процентилей)
        if use_files_count == "one":
            # Для одного файла используем только current_df
            manager_tb_mapping = build_manager_tb_mapping(current_df, pd.DataFrame())
            manager_gosb_mapping = build_manager_gosb_mapping(current_df, pd.DataFrame())
        elif use_files_count == "new":
            # Для нового варианта используем первый доступный файл 2025
            if not current_df.empty:
                manager_tb_mapping = build_manager_tb_mapping(current_df, pd.DataFrame())
                manager_gosb_mapping = build_manager_gosb_mapping(current_df, pd.DataFrame())
            else:
                # Если нет данных, возвращаем пустые Series
                manager_tb_mapping = pd.Series(dtype=object, name="tb")
                manager_gosb_mapping = pd.Series(dtype=object, name="gosb")
        else:
            manager_tb_mapping = build_manager_tb_mapping(current_df, previous_df)
            manager_gosb_mapping = build_manager_gosb_mapping(current_df, previous_df)
        
        # Добавляем ТБ и ГОСБ к summary_tn_combined
        summary_tn_combined["ТБ"] = summary_tn_combined[SELECTED_MANAGER_ID_COL].map(manager_tb_mapping).fillna("")
        summary_tn_combined["ГОСБ"] = summary_tn_combined[SELECTED_MANAGER_ID_COL].map(manager_gosb_mapping).fillna("")
        
        # Инициализируем калькулятор процентилей
        percentile_calc = PercentileCalculator()
        
        # Получаем параметры расчета процентиля
        percentile_type = percentile_calc_config.get("percentile_type", "above")
        percentile_group_by = percentile_calc_config.get("percentile_group_by", "all")
        percentile_filter = percentile_calc_config.get("percentile_filter", "all")
        
        log_info(logger, f"Параметры расчета процентиля: percentile_type={percentile_type}, percentile_group_by={percentile_group_by}, percentile_filter={percentile_filter}")
        
        # Определяем колонки для группировки
        percentile_tb_column = None
        percentile_gosb_column = None
        
        if percentile_group_by in ("tb", "tb_and_gosb"):
            percentile_tb_column = "ТБ"
        if percentile_group_by in ("gosb", "tb_and_gosb"):
            percentile_gosb_column = "ГОСБ"
        
        # Добавляем процентили для выбранного варианта
        # Используем summary_tn_combined, который уже содержит ТБ и ГОСБ
        selected_percentile = percentile_calc.append_percentile_columns(
            summary_tn_combined,
            value_column=value_column,  # Используем правильную колонку в зависимости от типа расчета
            tb_column=percentile_tb_column,
            gosb_column=percentile_gosb_column,
            percentile_filter=percentile_filter,
            group_by=percentile_group_by,
        )
        
        # Добавляем процентильные колонки
        percentile_columns = [col for col in selected_percentile.columns if col not in summary_tn_combined.columns]
        for col in percentile_columns:
            summary_tn_combined[col] = selected_percentile[col]
        
        # Переупорядочиваем колонки: сначала расчеты, потом процентили
        if use_files_count == "one":
            # Для одного файла базовые колонки зависят от типа расчета
            base_columns = [SELECTED_MANAGER_ID_COL, SELECTED_MANAGER_NAME_COL, "ТБ", "ГОСБ", value_column]
        elif use_files_count == "new":
            # Для нового варианта базовые колонки: сумма 2024, сумма 2025, месяцы
            base_columns = [SELECTED_MANAGER_ID_COL, SELECTED_MANAGER_NAME_COL, "ТБ", "ГОСБ",
                           "Сумма_2024", "Сумма_2025", "Месяцев_с_суммой_2024", "Месяцев_с_суммой_2025"]
        else:
            base_columns = [SELECTED_MANAGER_ID_COL, SELECTED_MANAGER_NAME_COL, "ТБ", "ГОСБ", 
                           "Факт_T0", "Факт_T1"]
            # Добавляем Факт_T2, если он есть (для варианта three)
            if "Факт_T2" in summary_tn_combined.columns:
                base_columns.append("Факт_T2")
            base_columns.append("Прирост")
            if "Количество записей" in summary_tn_combined.columns:
                base_columns.append("Количество записей")
        percentile_cols = [col for col in percentile_columns if col not in base_columns]
        summary_tn_combined = summary_tn_combined[base_columns + percentile_cols]
        
        # summary_tn уже содержит процентили (summary_tn_combined)
        # Используем его для всех листов
        summary_tn = summary_tn_combined.copy()
        percentile_tn = summary_tn_combined.copy()  # Оба содержат процентили
        
        # Создаём свод по ИНН для вариантов 2 и 3 (где key_mode="client" и use_files_count не "one" и не "new")
        client_summary_inn = None
        if use_files_count not in ["one", "new"] and key_mode == "client" and variant_df_for_client_summary is not None:
            client_summary_inn = build_client_summary_by_inn(
                variant_df=variant_df_for_client_summary,
                current_df=current_df,
                previous_df=previous_df,
                previous2_df=previous2_df if use_t2 else None,
                manager_tb_mapping=manager_tb_mapping,
                manager_gosb_mapping=manager_gosb_mapping,
                defaults=defaults,
                identifiers=identifiers,
                logger=logger,
            )
            log_info(logger, f"Создан свод по ИНН: {len(client_summary_inn)} клиентов")
        
        # Подготавливаем данные для SPOD
        spod_variants_config = spod_config.get("variants", [])
        spod_datasets: List[pd.DataFrame] = []
        csv_frames: List[pd.DataFrame] = []
        
        # Создаём маппинги ТБ и ГОСБ для менеджеров (уже созданы выше)
        
        # Обрабатываем каждый вариант SPOD
        for spod_variant in spod_variants_config:
            variant_name = spod_variant.get("name", "")
            source_type = spod_variant.get("source_type", "scenario_summary")
            calc_sheet_name = spod_variant.get("calc_sheet_name", "")
            
            # Определяем исходную таблицу
            # summary_tn уже содержит процентили, поэтому используем его для обоих вариантов
            source_table = percentile_tn  # Используем percentile_tn (который содержит все данные с процентилями)
            
            # Создаём SPOD датасет
            if should_write(variant_name, spod_variant_whitelist, "spod_variants"):
                # Определяем колонку для FACT_VALUE (для процентильного SPOD может отличаться от value_column)
                percentile_value_column = None
                
                # Для scenario_percentile всегда используем percentile_type из настроек процентиля
                if source_type == "scenario_percentile":
                    percentile_value_type = percentile_type
                    log_debug(
                        logger,
                        f"SPOD '{variant_name}': используется percentile_type из варианта процентиля: '{percentile_value_type}'",
                        class_name="ProjectProcessor",
                        func_name="process_project",
                    )
                else:
                    # Для scenario_summary percentile_value_type не используется
                    percentile_value_type = None
                
                if percentile_value_type:
                    if percentile_value_type == "above":
                        percentile_value_column = "Обогнал_всего_%"
                    elif percentile_value_type == "below":
                        percentile_value_column = "Обогнали_меня_всего_%"
                    else:
                        log_debug(
                            logger,
                            f"SPOD '{variant_name}': неизвестный percentile_value_type '{percentile_value_type}', используется value_column",
                            class_name="ProjectProcessor",
                            func_name="process_project",
                        )
                
                # Базовый SPOD датасет для CSV
                # Для режима "new" используем value_column из основного расчета, иначе из конфигурации SPOD варианта
                spod_value_column = value_column if use_files_count == "new" else spod_variant.get("value_column", "Прирост")
                spod_dataset = build_spod_dataset(
                    source_table=source_table,
                    value_column=spod_value_column,
                    fact_value_filter=spod_variant.get("fact_value_filter", ">0"),
                    plan_value=spod_variant.get("plan_value", 0.0),
                    priority=spod_variant.get("priority", 1),
                    contest_code=spod_variant.get("contest_code", ""),
                    tournament_code=spod_variant.get("tournament_code", ""),
                    contest_date=spod_variant.get("contest_date", "01/01/2025"),
                    identifiers=identifiers,
                    logger=logger,
                    dataset_name=variant_name,
                    percentile_value_column=percentile_value_column,
                )
                
                # Получаем отфильтрованную таблицу для добавления доп данных
                mask = build_filter_mask(source_table[spod_value_column], 
                                        spod_variant.get("fact_value_filter", ">0"))
                filtered_table = source_table[mask].copy()
                
                # Расширенный SPOD датасет для Excel (с дополнительными колонками)
                # Используем percentile_value_column для колонки "Факт", если она определена
                fact_column_for_excel = percentile_value_column if percentile_value_column is not None else spod_value_column
                spod_dataset_excel = build_spod_dataset_for_excel(
                    source_table=source_table,
                    filtered_table=filtered_table,
                    spod_dataset=spod_dataset,
                    value_column=fact_column_for_excel,  # Используем колонку для FACT_VALUE
                    source_type=source_type,  # Передаем тип источника для определения, нужны ли процентили
                    manager_tb_mapping=manager_tb_mapping,
                    manager_gosb_mapping=manager_gosb_mapping,
                    variant_df_for_client_summary=variant_df_for_client_summary if key_mode == "client" else None,
                    current_df=current_df if key_mode == "manager" else None,  # Для варианта 1 (по КМ)
                    previous_df=previous_df if key_mode == "manager" else None,  # Для варианта 1 (объединяем T-0 и T-1)
                    identifiers=identifiers,
                    logger=logger,
                )
                
                spod_datasets.append((variant_name, spod_dataset_excel))
                
                # Добавляем в CSV базовую версию (без доп данных), если указано
                if spod_variant.get("include_in_csv", False):
                    csv_frames.append(spod_dataset)
        
        # Подготавливаем таблицы для вывода
        raw_tables = {
            "RAW_T0": format_raw_sheet(current_df, current_alias_to_source),
            "RAW_T1": format_raw_sheet(previous_df, previous_alias_to_source),
        }
        if use_t2 and previous2_df is not None:
            previous2_column_profiles = build_column_profiles(get_file_columns(file_section, "previous2", defaults))
            previous2_alias_to_source = previous2_column_profiles["alias_to_source"]
            raw_tables["RAW_T2"] = format_raw_sheet(previous2_df, previous2_alias_to_source)


        report_suffix = timestamp_suffix()
        excel_name = f"{spod_config['file_prefix']}{report_suffix}.xlsx"
        excel_path = output_dir / excel_name
        log_info(logger, f"Сохраняю Excel-файл {excel_name}")

        log_info(
            logger,
            "Используется движок openpyxl (доступен в базовой поставке Anaconda) для сохранения отчёта.",
        )

        # Инициализируем экспортер Excel
        excel_exporter = ExcelExporter()
        
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            written_sheets: Set[str] = set()

            def write_sheet(sheet_name: str, table: pd.DataFrame) -> None:
                """Внутренняя функция для записи листа с проверкой дубликатов и сортировкой."""
                if sheet_name in written_sheets:
                    log_debug(
                        logger,
                        f"Лист {sheet_name} уже создан — пропускаю повторную запись",
                        class_name="ProjectProcessor",
                        func_name="process_project",
                    )
                    return
                
                # Сортируем таблицу в зависимости от типа листа (от большего к меньшему)
                table_to_write = table.copy()
                sort_column = None
                
                if sheet_name == "SUMMARY_TN":
                    # Для одного файла используем value_column, для нового варианта - "Сумма_2025", для двух/трех - "Прирост"
                    if use_files_count == "one":
                        sort_column = value_column
                    elif use_files_count == "new":
                        sort_column = "Сумма_2025"
                    else:
                        sort_column = "Прирост"
                elif sheet_name == "SUMMARY_INN":
                    sort_column = "Прирост"
                elif sheet_name in ["SPOD_SCENARIO", "SPOD_SCENARIO_PERCENTILE"]:
                    sort_column = "Факт"
                elif sheet_name in ["RAW_T0", "RAW_T1", "RAW_T2"]:
                    # Для RAW листов ищем колонку с фактом
                    if "Факт (число)" in table_to_write.columns:
                        sort_column = "Факт (число)"
                    elif "fact_value_clean" in table_to_write.columns:
                        sort_column = "fact_value_clean"
                
                if sort_column and sort_column in table_to_write.columns:
                    table_to_write = table_to_write.sort_values(
                        by=sort_column,
                        ascending=False,
                        na_position="last"
                    )
                    log_debug(
                        logger,
                        f"Лист {sheet_name}: отсортирован по {sort_column} (убывание)",
                        class_name="ProjectProcessor",
                        func_name="process_project",
                    )
                
                excel_exporter.write_sheet(
                    writer, 
                    sheet_name, 
                    table_to_write, 
                    written_sheets,
                    min_width=min_width,
                    max_width=max_width,
                    wrap_text=wrap_text,
                )

            # Записываем SUMMARY_TN (объединенный с процентилями)
            if should_write("SUMMARY_TN", summary_sheet_whitelist, "summary_sheets"):
                write_sheet("SUMMARY_TN", percentile_tn)
            
            # Записываем свод по ИНН для вариантов 2 и 3
            if client_summary_inn is not None:
                if should_write("SUMMARY_INN", summary_sheet_whitelist, "summary_sheets"):
                    write_sheet("SUMMARY_INN", client_summary_inn)
            
            # Записываем SPOD листы
            for variant_name, spod_dataset in spod_datasets:
                if should_write(variant_name, spod_variant_whitelist, "spod_variants"):
                    write_sheet(variant_name, spod_dataset)

            # Записываем raw таблицы
            for sheet_name, raw_table in raw_tables.items():
                if not should_write(sheet_name, raw_sheet_whitelist, "raw_sheets"):
                    continue
                write_sheet(sheet_name, raw_table)
        
        # Создаём CSV файл, если есть данные для выгрузки
        if csv_frames:
            csv_name = f"{spod_config['file_prefix']}_SPOD{report_suffix}.csv"
            csv_path = output_dir / csv_name
            log_info(logger, f"Сохраняю CSV-файл {csv_name}")
            
            combined_csv = pd.concat(csv_frames, ignore_index=True)
            combined_csv.to_csv(
                csv_path,
                sep=";",
                index=False,
                quoting=csv.QUOTE_MINIMAL,
                encoding="utf-8-sig",  # UTF-8 с BOM для корректного отображения в Excel
            )
            log_info(logger, f"CSV-файл сохранён: {csv_name} ({len(combined_csv)} строк)")

        log_info(logger, "Обработка успешно завершена")
    except Exception as exc:
        log_info(logger, f"Обработка завершилась с ошибкой: {exc}")
        stack = traceback.format_exc().replace("\n", " | ")
        log_debug(
            logger,
            f"Трассировка: {stack}",
            class_name="Main",
            func_name="process_project",
        )
        raise


def main() -> None:
    """Точка входа."""

    project_root = Path(__file__).resolve().parent.parent
    process_project(project_root)


if __name__ == "__main__":
    main()
